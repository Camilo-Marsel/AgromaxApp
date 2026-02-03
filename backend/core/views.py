# backend/core/views.py

from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from .services.nomina_calculator import NominaCalculator
from django.http import FileResponse
from .services.pdf_generator import generar_comprobante_pdf
from .services.excel_generator import ExcelGenerator
from rest_framework.filters import SearchFilter 
from django.utils import timezone


from .models import (
    Usuario, Rol, TipoContrato, Finca, Lote, Trabajador,
    UnidadMedida, Labor, ListaPrecios, VariablesNomina,
    Quincena, RegistroLabor, Nomina, DetalleNomina,
    Prestamo, CuotaPrestamo, AuditoriaLog,
    Contrato, DocumentoContrato, ConfiguracionEmpresa,
    PILA, DetallePILA, ProvisionPrestaciones, ResumenPrestaciones,
    PORCENTAJE_SALUD_EMPLEADOR, PORCENTAJE_PENSION_EMPLEADOR,
    PORCENTAJE_ARL, PORCENTAJE_CAJA_COMPENSACION,
    PORCENTAJE_CESANTIAS, PORCENTAJE_INTERESES_CESANTIAS,
    PORCENTAJE_PRIMA, PORCENTAJE_VACACIONES,
)
from .serializers import *
from .permissions import IsAdministrador, IsSupervisorOrAbove, CanModifyData, ReadOnly, FincaFilterMixin

# Alias para mantener compatibilidad
IsSuperAdmin = IsAdministrador
IsDigitadorOrAbove = IsSupervisorOrAbove
from .filters import *

# ============================================================================
# USUARIOS Y ROLES
# ============================================================================

class RolViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet de solo lectura para Roles"""
    queryset = Rol.objects.all()
    serializer_class = RolSerializer
    permission_classes = [permissions.IsAuthenticated]


class UsuarioViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de usuarios"""
    queryset = Usuario.objects.all()
    permission_classes = [IsSuperAdmin]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['username', 'email', 'first_name', 'last_name']
    ordering_fields = ['username', 'date_joined']
    ordering = ['-date_joined']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return UsuarioCreateSerializer
        return UsuarioSerializer


# ============================================================================
# TRABAJADORES
# ============================================================================

class TipoContratoViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet de solo lectura para Tipos de Contrato"""
    queryset = TipoContrato.objects.all()
    serializer_class = TipoContratoSerializer
    permission_classes = [permissions.IsAuthenticated]


class TrabajadorViewSet(FincaFilterMixin, viewsets.ModelViewSet):
    """ViewSet para gestión de trabajadores"""
    queryset = Trabajador.objects.all()
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = TrabajadorFilter
    search_fields = ['nombres', 'apellidos', 'numero_documento']
    ordering_fields = ['apellidos', 'fecha_ingreso', 'created_at']
    ordering = ['apellidos', 'nombres']
    finca_field = 'finca'  # Filtrar por finca del trabajador
    
    def get_serializer_class(self):
        if self.action == 'list':
            return TrabajadorListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return TrabajadorCreateUpdateSerializer
        return TrabajadorDetailSerializer
    
    @action(detail=True, methods=['post'])
    def retirar(self, request, pk=None):
        """
        Retirar un trabajador.
        Requiere: motivo_retiro (y opcionalmente observaciones_retiro, fecha_retiro)
        """
        from django.utils import timezone

        trabajador = self.get_object()

        if trabajador.estado == Trabajador.RETIRADO:
            return Response(
                {'error': 'El trabajador ya está retirado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar motivo de retiro
        motivo = request.data.get('motivo_retiro')
        if not motivo:
            return Response(
                {'error': 'Debe indicar el motivo de retiro'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar si tiene adelantos pendientes
        adelantos_pendientes = trabajador.prestamos.filter(estado='ACTIVO')
        if adelantos_pendientes.exists():
            total_pendiente = sum(a.saldo_pendiente for a in adelantos_pendientes)
            # Solo advertir, no bloquear
            warning = f'El trabajador tiene adelantos pendientes por ${total_pendiente}'
        else:
            warning = None

        # Actualizar trabajador
        trabajador.estado = Trabajador.RETIRADO
        trabajador.motivo_retiro = motivo
        trabajador.observaciones_retiro = request.data.get('observaciones_retiro', '')
        trabajador.fecha_retiro = request.data.get('fecha_retiro') or timezone.now().date()
        trabajador.save()

        serializer = self.get_serializer(trabajador)
        response_data = serializer.data
        if warning:
            response_data['warning'] = warning

        return Response(response_data)

    @action(detail=True, methods=['post'])
    def reactivar(self, request, pk=None):
        """
        Reactivar un trabajador retirado.
        El trabajador vuelve a estado SIN_CONTRATO.
        """
        trabajador = self.get_object()

        if trabajador.estado != Trabajador.RETIRADO:
            return Response(
                {'error': 'Solo se pueden reactivar trabajadores retirados'},
                status=status.HTTP_400_BAD_REQUEST
            )

        trabajador.estado = Trabajador.SIN_CONTRATO
        # Limpiar campos de retiro
        trabajador.fecha_retiro = None
        trabajador.motivo_retiro = None
        trabajador.observaciones_retiro = None
        trabajador.save()

        serializer = self.get_serializer(trabajador)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='lista-simple')
    def lista_simple(self, request):
        """Lista simple sin paginación (útil para selects en formularios)"""
        qs = self.filter_queryset(self.get_queryset())
        serializer = TrabajadorListSerializer(qs, many=True, context={'request': request})
        return Response(serializer.data)



# ============================================================================
# FINCAS Y LOTES
# ============================================================================

class FincaViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de fincas"""
    queryset = Finca.objects.all()
    serializer_class = FincaSerializer
    permission_classes = [CanModifyData]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['nombre', 'ubicacion']
    ordering_fields = ['nombre', 'created_at']
    ordering = ['nombre']

    def get_queryset(self):
        """Filtra fincas por asignación del usuario"""
        queryset = super().get_queryset()
        user = self.request.user

        if not user or not user.is_authenticated:
            return queryset.none()

        # Admins ven todas las fincas
        if user.es_administrador:
            return queryset

        # Otros usuarios solo ven sus fincas asignadas
        return user.fincas_asignadas.all()

    @action(detail=True, methods=['get'])
    def trabajadores(self, request, pk=None):
        """Obtener trabajadores activos de una finca (CONTRATADO o SIN_CONTRATO)"""
        finca = self.get_object()
        # Excluir trabajadores retirados
        trabajadores = finca.trabajadores.exclude(estado=Trabajador.RETIRADO)
        serializer = TrabajadorListSerializer(trabajadores, many=True, context={'request': request})
        return Response(serializer.data)


class LoteViewSet(FincaFilterMixin, viewsets.ModelViewSet):
    """ViewSet para gestión de lotes"""
    queryset = Lote.objects.all()
    serializer_class = LoteSerializer
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    finca_field = 'finca'  # Filtrar por finca
    filterset_fields = ['finca', 'activo']
    search_fields = ['nombre']
    ordering_fields = ['nombre', 'finca__nombre']
    ordering = ['finca__nombre', 'nombre']


# ============================================================================
# CATÁLOGOS
# ============================================================================

class UnidadMedidaViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet de solo lectura para Unidades de Medida"""
    queryset = UnidadMedida.objects.all()
    serializer_class = UnidadMedidaSerializer
    permission_classes = [permissions.IsAuthenticated]


class LaborViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de labores"""
    queryset = Labor.objects.all()
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = LaborFilter
    search_fields = ['codigo', 'nombre']
    ordering_fields = ['codigo', 'nombre', 'created_at']
    ordering = ['nombre']
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return LaborCreateUpdateSerializer
        return LaborListSerializer


class ListaPreciosViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de precios"""
    queryset = ListaPrecios.objects.all()
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['labor', 'fecha_inicio_vigencia']
    ordering_fields = ['fecha_inicio_vigencia', 'created_at']
    ordering = ['-fecha_inicio_vigencia']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ListaPreciosCreateSerializer
        return ListaPreciosSerializer
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class VariablesNominaViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de variables de nómina"""
    queryset = VariablesNomina.objects.all()
    serializer_class = VariablesNominaSerializer
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['nombre']
    ordering = ['-fecha_inicio_vigencia']
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


# ============================================================================
# QUINCENAS Y REGISTROS
# ============================================================================

class QuincenaViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de quincenas"""
    queryset = Quincena.objects.all()
    serializer_class = QuincenaSerializer
    permission_classes = [CanModifyData]
    filter_backends = [OrderingFilter]
    ordering = ['-año', '-mes', '-numero']
    
    @action(detail=False, methods=['post'])
    def crear_actual(self, request):
        """Crear quincena actual si no existe"""
        from django.utils import timezone
        from datetime import datetime, timedelta
        
        hoy = timezone.now().date()
        año = hoy.year
        mes = hoy.month
        
        # Determinar si es primera o segunda quincena
        if hoy.day <= 15:
            numero = 1
            fecha_inicio = hoy.replace(day=1)
            fecha_fin = hoy.replace(day=15)
        else:
            numero = 2
            fecha_inicio = hoy.replace(day=16)
            # Último día del mes
            if mes == 12:
                fecha_fin = hoy.replace(day=31)
            else:
                fecha_fin = (hoy.replace(month=mes + 1, day=1) - timedelta(days=1))
        
        fecha_cierre_registro = fecha_fin + timedelta(days=15)
        
        # Verificar si ya existe
        quincena, created = Quincena.objects.get_or_create(
            año=año,
            mes=mes,
            numero=numero,
            defaults={
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
                'fecha_cierre_registro': fecha_cierre_registro,
                'estado': 'ABIERTA'
            }
        )
        
        serializer = self.get_serializer(quincena)
        
        if created:
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(
                {'message': 'La quincena actual ya existe', 'data': serializer.data},
                status=status.HTTP_200_OK
            )
    
    @action(detail=True, methods=['post'])
    def cerrar(self, request, pk=None):
        """Cerrar quincena para preparar nómina"""
        quincena = self.get_object()
        quincena.estado = 'EN_CALCULO'
        quincena.save()
        serializer = self.get_serializer(quincena)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def estadisticas(self, request, pk=None):
        """Obtener estadísticas de registros de la quincena"""
        quincena = self.get_object()

        # Total de trabajadores que pueden trabajar (no retirados)
        trabajadores_activos = Trabajador.objects.exclude(estado=Trabajador.RETIRADO).count()

        # Trabajadores con al menos un registro en esta quincena
        trabajadores_con_registros = RegistroLabor.objects.filter(
            quincena=quincena
        ).values('trabajador').distinct().count()

        # Total de registros
        total_registros = RegistroLabor.objects.filter(quincena=quincena).count()

        # Trabajadores sin registros
        trabajadores_sin_registros = trabajadores_activos - trabajadores_con_registros

        return Response({
            'trabajadores_activos': trabajadores_activos,
            'trabajadores_con_registros': trabajadores_con_registros,
            'trabajadores_sin_registros': trabajadores_sin_registros,
            'total_registros': total_registros,
            'porcentaje_cobertura': round(
                (trabajadores_con_registros / trabajadores_activos * 100) if trabajadores_activos > 0 else 0,
                2
            )
        })


class RegistroLaborViewSet(FincaFilterMixin, viewsets.ModelViewSet):
    """ViewSet para gestión de registros de labores"""
    queryset = RegistroLabor.objects.all()
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = RegistroLaborFilter
    ordering = ['-fecha', 'trabajador']
    finca_field = 'trabajador__finca'  # Filtrar por finca del trabajador
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return RegistroLaborCreateUpdateSerializer
        return RegistroLaborSerializer
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
    
    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['post'])
    def crear_multiples(self, request):
        """Crear múltiples registros de labor (para labores tipo DÍA con múltiples fechas).

        Soporta dos modos:
        - Sin cantidad_total: cada registro tiene cantidad=1 (labores tipo DÍA)
        - Con cantidad_total: la cantidad se reparte equitativamente entre los días válidos
          (ej: Desmache con 3.5 hectáreas en 5 días = 0.7 por día)
        """
        trabajador_id = request.data.get('trabajador')
        labor_id = request.data.get('labor')
        fechas = request.data.get('fechas', [])  # Lista de fechas
        quincena_id = request.data.get('quincena')
        observaciones = request.data.get('observaciones', '')
        cantidad_total = request.data.get('cantidad_total')  # Opcional: cantidad total a repartir

        if not all([trabajador_id, labor_id, fechas, quincena_id]):
            return Response(
                {'error': 'Faltan campos requeridos'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            trabajador = Trabajador.objects.get(id=trabajador_id)
            labor = Labor.objects.get(id=labor_id)
            quincena = Quincena.objects.get(id=quincena_id)
        except (Trabajador.DoesNotExist, Labor.DoesNotExist, Quincena.DoesNotExist):
            return Response(
                {'error': 'Trabajador, Labor o Quincena no encontrados'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Primero validar todas las fechas para saber cuántas son válidas
        from datetime import datetime as dt
        from decimal import Decimal, ROUND_DOWN

        fechas_validas = []
        errores = []

        for fecha_str in fechas:
            fecha = dt.strptime(fecha_str, '%Y-%m-%d').date()

            if fecha.weekday() == 6:
                errores.append(f"{fecha_str}: No se puede registrar en domingo")
                continue

            if RegistroLabor.objects.filter(
                trabajador=trabajador,
                fecha=fecha,
                quincena=quincena
            ).exists():
                errores.append(f"{fecha_str}: Ya existe un registro para este día")
                continue

            fechas_validas.append(fecha)

        if not fechas_validas:
            return Response({
                'message': 'No se crearon registros',
                'registros': [],
                'errores': errores
            }, status=status.HTTP_400_BAD_REQUEST)

        # Calcular cantidad por día
        if cantidad_total is not None:
            total = Decimal(str(cantidad_total))
            n = len(fechas_validas)
            cantidad_por_dia = (total / n).quantize(Decimal('0.0001'), rounding=ROUND_DOWN)
            # Ajustar residuo en el último día para que sume exactamente
            residuo = total - (cantidad_por_dia * n)
        else:
            cantidad_por_dia = Decimal('1')
            residuo = Decimal('0')

        # Crear registros
        registros_creados = []
        for i, fecha in enumerate(fechas_validas):
            cantidad = cantidad_por_dia
            if i == len(fechas_validas) - 1:
                cantidad += residuo  # Último día absorbe el residuo

            registro = RegistroLabor.objects.create(
                trabajador=trabajador,
                labor=labor,
                fecha=fecha,
                quincena=quincena,
                cantidad=cantidad,
                observaciones=observaciones,
                created_by=request.user
            )
            registros_creados.append(registro)

        serializer = self.get_serializer(registros_creados, many=True)

        return Response({
            'message': f'{len(registros_creados)} registros creados correctamente',
            'registros': serializer.data,
            'errores': errores
        }, status=status.HTTP_201_CREATED)

# ============================================================================
# NÓMINA
# ============================================================================




class NominaViewSet(FincaFilterMixin, viewsets.ModelViewSet):
    """ViewSet para gestión de nóminas"""
    queryset = Nomina.objects.all()
    serializer_class = NominaSerializer
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = NominaFilter
    ordering = ['-quincena', 'trabajador']
    finca_field = 'trabajador__finca'  # Filtrar por finca del trabajador
    
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    
    filterset_class = NominaFilter
    search_fields = ['trabajador__nombres', 'trabajador__apellidos', 'trabajador__numero_documento']
    
    ordering = ['-quincena', 'trabajador']
    
    # AGREGAR esto para desactivar paginación (mostrar todas):
    pagination_class = None

    def get_queryset(self):
        return super().get_queryset().exclude(
            trabajador__estado=Trabajador.RETIRADO
        )

    @action(detail=False, methods=['get'])
    def exportar_excel_quincena(self, request):
        """Exportar lista de pagos de una quincena a Excel"""
        quincena_id = request.query_params.get('quincena')
        
        if not quincena_id:
            return Response(
                {'error': 'Se requiere el parámetro quincena'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            quincena = Quincena.objects.get(id=quincena_id)
        except Quincena.DoesNotExist:
            return Response(
                {'error': 'Quincena no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Obtener nóminas de la quincena, excluyendo retirados
        nominas = self.queryset.filter(
            quincena=quincena,
            estado__in=['PENDIENTE', 'APROBADA']
        ).exclude(
            trabajador__estado='RETIRADO'
        ).select_related('trabajador', 'quincena')

        # Filtro de fincas: soporta multi-finca (fincas=1,2,3) y single (finca=1)
        fincas_param = request.query_params.get('fincas', '')
        finca_id = request.query_params.get('finca', '')

        if fincas_param:
            fincas_ids = [int(f) for f in fincas_param.split(',') if f.strip()]
            nominas = nominas.filter(trabajador__finca__id__in=fincas_ids)
        elif finca_id:
            nominas = nominas.filter(trabajador__finca__id=finca_id)

        nominas = nominas.order_by('trabajador__apellidos', 'trabajador__nombres')

        if not nominas.exists():
            return Response(
                {'error': 'No hay nóminas calculadas para esta quincena'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generar Excel
        generator = ExcelGenerator(nominas, quincena)
        return generator.generar_lista_pagos()

    @action(detail=False, methods=['post'])
    def calcular_quincena(self, request):
        """Calcular nómina para toda la quincena"""
        quincena_id = request.data.get('quincena_id')
        
        if not quincena_id:
            return Response(
                {'error': 'quincena_id es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            quincena = Quincena.objects.get(id=quincena_id)
        except Quincena.DoesNotExist:
            return Response(
                {'error': 'Quincena no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Calcular nómina
        calculator = NominaCalculator(quincena)
        nominas = calculator.calcular_quincena_completa(request.user)
        
        serializer = self.get_serializer(nominas, many=True)
        
        return Response({
            'message': f'Nómina calculada para {len(nominas)} trabajadores',
            'nominas': serializer.data
        })
    

    @action(detail=True, methods=['post'])
    def recalcular(self, request, pk=None):
        """Recalcular nómina manteniendo ajustes manuales"""
        nomina = self.get_object()
        
        # Solo se puede recalcular si está en PENDIENTE
        if nomina.estado != 'PENDIENTE':
            return Response(
                {'error': 'Solo se pueden recalcular nóminas en estado PENDIENTE'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Importar calculador
            from .services.nomina_calculator import NominaCalculator
            
            # Guardar ajustes manuales actuales
            devengos_adicionales = nomina.devengos_adicionales
            descripcion_devengos = nomina.descripcion_devengos_adicionales
            deducciones_adicionales = nomina.deducciones_adicionales
            descripcion_deducciones = nomina.descripcion_deducciones_adicionales
            
            # Recalcular
            calculator = NominaCalculator(nomina.quincena)
            nomina_actualizada = calculator.calcular_trabajador(
                nomina.trabajador,
                request.user
            )
            
            # Restaurar ajustes manuales
            nomina_actualizada.devengos_adicionales = devengos_adicionales
            nomina_actualizada.descripcion_devengos_adicionales = descripcion_devengos
            nomina_actualizada.deducciones_adicionales = deducciones_adicionales
            nomina_actualizada.descripcion_deducciones_adicionales = descripcion_deducciones
            nomina_actualizada.save()
            
            # Recalcular con ajustes
            nomina_actualizada = calculator.calcular_trabajador(
                nomina.trabajador,
                request.user
            )
            
            serializer = self.get_serializer(nomina_actualizada)
            return Response(serializer.data)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def aprobar(self, request, pk=None):
        """Aprobar nómina pendiente"""
        nomina = self.get_object()

        if nomina.estado != 'PENDIENTE':
            return Response(
                {'error': 'Solo se pueden aprobar nóminas en estado PENDIENTE'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Cambiar estado
        nomina.estado = 'APROBADA'
        nomina.fecha_aprobacion = timezone.now()
        nomina.save()

        serializer = self.get_serializer(nomina)
        return Response(serializer.data)


    @action(detail=True, methods=['post'])
    def rechazar(self, request, pk=None):
        """Rechazar nómina aprobada (regresa a PENDIENTE)"""
        nomina = self.get_object()

        if nomina.estado != 'APROBADA':
            return Response(
                {'error': 'Solo se pueden rechazar nóminas en estado APROBADA'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener motivo del rechazo (opcional)
        motivo = request.data.get('motivo', '')

        # Regresar a pendiente
        nomina.estado = 'PENDIENTE'
        nomina.fecha_aprobacion = None

        # Agregar motivo a observaciones
        if motivo:
            fecha_hora = timezone.now().strftime('%Y-%m-%d %H:%M')
            nota = f"\n[{fecha_hora}] Nómina rechazada: {motivo}"
            nomina.observaciones = (nomina.observaciones or '') + nota

        nomina.save()

        serializer = self.get_serializer(nomina)
        return Response(serializer.data)


    @action(detail=False, methods=['post'])
    def aprobar_masivo(self, request):
        """Aprobar todas las nóminas PENDIENTES de una quincena (con filtro opcional de finca)"""
        quincena_id = request.data.get('quincena_id')
        finca_id = request.data.get('finca_id')

        if not quincena_id:
            return Response(
                {'error': 'El ID de quincena es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Filtrar nóminas
        nominas = Nomina.objects.filter(
            quincena_id=quincena_id,
            estado='PENDIENTE'
        )

        # Aplicar filtro de finca si se proporciona
        if finca_id:
            nominas = nominas.filter(trabajador__finca_id=finca_id)

        # Contar antes de actualizar
        count = nominas.count()

        if count == 0:
            return Response(
                {'error': 'No hay nóminas PENDIENTES para aprobar con los filtros aplicados'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Aprobar todas
        nominas.update(
            estado='APROBADA',
            fecha_aprobacion=timezone.now()
        )

        return Response({
            'message': f'{count} nómina(s) aprobada(s) correctamente',
            'count': count
        })


    @action(detail=False, methods=['post'])
    def rechazar_masivo(self, request):
        """Rechazar todas las nóminas APROBADAS de una quincena (regresa a PENDIENTE)"""
        quincena_id = request.data.get('quincena_id')
        finca_id = request.data.get('finca_id')
        motivo = request.data.get('motivo', '')

        if not quincena_id:
            return Response(
                {'error': 'El ID de quincena es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Filtrar nóminas
        nominas = Nomina.objects.filter(
            quincena_id=quincena_id,
            estado='APROBADA'
        )

        # Aplicar filtro de finca si se proporciona
        if finca_id:
            nominas = nominas.filter(trabajador__finca_id=finca_id)

        # Contar antes de actualizar
        count = nominas.count()

        if count == 0:
            return Response(
                {'error': 'No hay nóminas APROBADAS para rechazar con los filtros aplicados'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Agregar motivo a observaciones si se proporciona
        if motivo:
            fecha_hora = timezone.now().strftime('%Y-%m-%d %H:%M')
            for nomina in nominas:
                nota = f"\n[{fecha_hora}] Rechazo masivo: {motivo}"
                nomina.observaciones = (nomina.observaciones or '') + nota
                nomina.estado = 'PENDIENTE'
                nomina.fecha_aprobacion = None
                nomina.save()
        else:
            nominas.update(
                estado='PENDIENTE',
                fecha_aprobacion=None
            )

        return Response({
            'message': f'{count} nómina(s) rechazada(s) correctamente',
            'count': count
        })
    
    @action(detail=True, methods=['get'])
    def descargar_pdf(self, request, pk=None):
        """Descargar comprobante de pago en PDF"""
        nomina = self.get_object()

        # Generar PDF
        pdf_buffer = generar_comprobante_pdf(nomina)

        # Nombre del archivo
        filename = f"comprobante_nomina_{nomina.trabajador.numero_documento}_{nomina.quincena.año}_{nomina.quincena.mes}_Q{nomina.quincena.numero}.pdf"

        # Retornar como descarga
        response = FileResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    @action(detail=False, methods=['get'])
    def descargar_colillas_consolidadas(self, request):
        """
        Descargar colillas consolidadas en un solo PDF.
        Combina todas las colillas de pago según los filtros aplicados.

        Query params:
        - quincena (required): ID de la quincena
        - finca (optional): ID de la finca
        - estados (optional): Estados separados por coma (ej: PENDIENTE,APROBADA)
        """
        from .services.pdf_generator import generar_colillas_consolidadas

        # Obtener parámetros
        quincena_id = request.query_params.get('quincena')
        finca_id = request.query_params.get('finca')
        estados_str = request.query_params.get('estados', '')

        if not quincena_id:
            return Response(
                {'error': 'El parámetro quincena es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que la quincena existe
        try:
            quincena = Quincena.objects.get(id=quincena_id)
        except Quincena.DoesNotExist:
            return Response(
                {'error': 'Quincena no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Construir filtro de nóminas
        filtros = {'quincena': quincena}

        if finca_id:
            filtros['trabajador__finca_id'] = finca_id

        # Filtrar por estados si se proporcionan
        if estados_str:
            estados = [e.strip() for e in estados_str.split(',') if e.strip()]
            if estados:
                filtros['estado__in'] = estados

        # Obtener nóminas ordenadas por trabajador
        nominas = Nomina.objects.filter(**filtros).select_related(
            'trabajador', 'quincena', 'trabajador__finca', 'trabajador__tipo_contrato'
        ).order_by('trabajador__nombre_completo')

        if not nominas.exists():
            return Response(
                {'error': 'No se encontraron nóminas con los filtros especificados'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Generar PDF consolidado
        pdf_buffer = generar_colillas_consolidadas(nominas)

        # Nombre del archivo
        finca_nombre = f"_{Finca.objects.get(id=finca_id).nombre}" if finca_id else ""
        filename = f"Colillas_Q{quincena.numero}_{quincena.año}{finca_nombre}.pdf"

        # Retornar como descarga
        response = FileResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    @action(detail=True, methods=['post'])
    def enviar_recibo(self, request, pk=None):
        """Enviar recibo de nómina por correo electrónico al trabajador"""
        nomina = self.get_object()

        # Verificar que la nómina esté en estado válido
        if nomina.estado not in ['PENDIENTE', 'APROBADA']:
            return Response(
                {'error': 'Solo se pueden enviar recibos de nóminas pendientes o aprobadas'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar que el trabajador tenga correo
        if not nomina.trabajador.correo:
            return Response(
                {'error': f'El trabajador {nomina.trabajador.nombre_completo} no tiene correo registrado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from .services.email_service import enviar_recibo_nomina

            resultado = enviar_recibo_nomina(nomina)

            if resultado['success']:
                return Response({
                    'message': resultado['message'],
                    'correo': nomina.trabajador.correo
                })
            else:
                return Response(
                    {'error': resultado['message']},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            return Response(
                {'error': f'Error al enviar correo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def enviar_recibos_masivo(self, request):
        """
        Enviar recibos de nómina por correo electrónico a múltiples trabajadores.

        Body params:
        - quincena_id (required): ID de la quincena
        - finca_id (optional): Filtrar por finca
        - estados (optional): Lista de estados (por defecto ['APROBADA'])
        """
        quincena_id = request.data.get('quincena_id')
        finca_id = request.data.get('finca_id')
        estados = request.data.get('estados', ['APROBADA'])

        if not quincena_id:
            return Response(
                {'error': 'El ID de quincena es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Filtrar nóminas
        nominas = Nomina.objects.filter(
            quincena_id=quincena_id,
            estado__in=estados,
            trabajador__correo__isnull=False
        ).exclude(trabajador__correo='').select_related('trabajador', 'quincena')

        # Aplicar filtro de finca si se proporciona
        if finca_id:
            nominas = nominas.filter(trabajador__finca_id=finca_id)

        if not nominas.exists():
            return Response(
                {'error': 'No hay nóminas con correo registrado para enviar'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from .services.email_service import enviar_recibos_masivo

            resultado = enviar_recibos_masivo(nominas)

            return Response({
                'message': f'Proceso completado: {resultado["exitosos"]} enviados, {resultado["fallidos"]} fallidos',
                'exitosos': resultado['exitosos'],
                'fallidos': resultado['fallidos'],
                'detalles': resultado['detalles']
            })
        except Exception as e:
            return Response(
                {'error': f'Error en el envío masivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ============================================================================
# AUDITORÍA
# ============================================================================

class AuditoriaLogViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet de solo lectura para logs de auditoría"""
    queryset = AuditoriaLog.objects.all()
    serializer_class = AuditoriaLogSerializer
    permission_classes = [IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['accion', 'tabla_afectada', 'usuario']
    ordering = ['-created_at']


# ============================================================================
# PRESTAMOS Y CUOTAS
# ============================================================================  

class PrestamoViewSet(FincaFilterMixin, viewsets.ModelViewSet):
    """ViewSet para gestión de préstamos (adelantos de nómina)"""
    queryset = Prestamo.objects.all()
    serializer_class = PrestamoSerializer
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['trabajador', 'estado', 'tipo_pago']
    search_fields = ['trabajador__nombres', 'trabajador__apellidos', 'trabajador__numero_documento']
    ordering = ['-fecha_prestamo']
    finca_field = 'trabajador__finca'  # Filtrar por finca del trabajador
    
    def get_serializer_class(self):
        if self.action == 'create':
            return PrestamoCreateSerializer
        return PrestamoSerializer
    
    @action(detail=True, methods=['post'])
    def cancelar(self, request, pk=None):
        """Cancelar un préstamo (marca como CANCELADO)"""
        prestamo = self.get_object()
        
        if prestamo.estado == 'PAGADO':
            return Response(
                {'error': 'No se puede cancelar un préstamo ya pagado'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        prestamo.estado = 'CANCELADO'
        prestamo.save()
        
        # Cancelar cuotas pendientes
        CuotaPrestamo.objects.filter(
            prestamo=prestamo,
            estado='PENDIENTE'
        ).update(estado='CANCELADA')
        
        serializer = self.get_serializer(prestamo)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def generar_autorizacion(self, request, pk=None):
        """Generar documento de autorización de descuento PDF"""
        prestamo = self.get_object()
        
        from .services.prestamo_pdf import generar_autorizacion_pdf
        
        # Generar PDF
        pdf_buffer = generar_autorizacion_pdf(prestamo)
        
        # Nombre del archivo
        filename = f"autorizacion_prestamo_{prestamo.id}_{prestamo.trabajador.numero_documento}.pdf"
        
        # Retornar como descarga
        response = FileResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @action(detail=True, methods=['get'])
    def generar_paz_y_salvo(self, request, pk=None):
        """Generar certificado de paz y salvo PDF"""
        prestamo = self.get_object()
        
        if prestamo.estado != 'PAGADO':
            return Response(
                {'error': 'Solo se puede generar paz y salvo para préstamos pagados'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        from .services.prestamo_pdf import generar_paz_y_salvo_pdf
        
        # Generar PDF
        pdf_buffer = generar_paz_y_salvo_pdf(prestamo)
        
        # Nombre del archivo
        filename = f"paz_y_salvo_prestamo_{prestamo.id}_{prestamo.trabajador.numero_documento}.pdf"
        
        # Retornar como descarga
        response = FileResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class VariablesNominaViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de variables de nómina"""
    queryset = VariablesNomina.objects.all()
    serializer_class = VariablesNominaSerializer
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['nombre']
    ordering = ['-fecha_inicio_vigencia']
    
    @action(detail=False, methods=['get'])
    def vigentes(self, request):
        """Obtener todas las variables (vigentes o no) para la página de configuración"""
        from django.utils import timezone
        hoy = timezone.now().date()

        # Nombres legibles para las variables
        NOMBRES_DISPLAY = {
            VariablesNomina.SALARIO_MINIMO: 'Salario Mínimo',
            VariablesNomina.AUXILIO_TRANSPORTE: 'Auxilio de Transporte',
            VariablesNomina.PORCENTAJE_SALUD: 'Porcentaje Salud',
            VariablesNomina.PORCENTAJE_PENSION: 'Porcentaje Pensión',
        }

        variables_resultado = {}

        for nombre_var in [
            VariablesNomina.SALARIO_MINIMO,
            VariablesNomina.AUXILIO_TRANSPORTE,
            VariablesNomina.PORCENTAJE_SALUD,
            VariablesNomina.PORCENTAJE_PENSION
        ]:
            # Buscar variable vigente
            variable = VariablesNomina.objects.filter(
                nombre=nombre_var,
                fecha_inicio_vigencia__lte=hoy,
                fecha_fin_vigencia__isnull=True
            ).first()

            if variable:
                variables_resultado[nombre_var] = {
                    'id': variable.id,
                    'nombre': variable.get_nombre_display(),
                    'valor': str(variable.valor),
                    'fecha_inicio_vigencia': variable.fecha_inicio_vigencia,
                    'tiene_vigente': True,
                }
            else:
                # Variable sin valor vigente - mostrar para que pueda crearse
                variables_resultado[nombre_var] = {
                    'id': None,
                    'nombre': NOMBRES_DISPLAY.get(nombre_var, nombre_var),
                    'valor': None,
                    'fecha_inicio_vigencia': None,
                    'tiene_vigente': False,
                }

        return Response(variables_resultado)
    
    @action(detail=False, methods=['post'])
    def actualizar_variable(self, request):
        """Crear nueva versión de una variable (cierra la anterior)"""
        nombre = request.data.get('nombre')
        valor = request.data.get('valor')
        fecha_inicio = request.data.get('fecha_inicio_vigencia')
        descripcion = request.data.get('descripcion', '')
        
        if not all([nombre, valor, fecha_inicio]):
            return Response(
                {'error': 'Faltan campos requeridos'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Cerrar variable anterior
            VariablesNomina.objects.filter(
                nombre=nombre,
                fecha_fin_vigencia__isnull=True
            ).update(fecha_fin_vigencia=fecha_inicio)
            
            # Crear nueva variable
            nueva_variable = VariablesNomina.objects.create(
                nombre=nombre,
                valor=valor,
                fecha_inicio_vigencia=fecha_inicio,
                descripcion=descripcion,
                created_by=request.user
            )
            
            serializer = self.get_serializer(nueva_variable)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )# Código temporal para agregar al final de views.py

# ============================================================================
# CONTRATOS
# ============================================================================

class ContratoViewSet(FincaFilterMixin, viewsets.ModelViewSet):
    """ViewSet para gestión de contratos laborales"""
    queryset = Contrato.objects.select_related('trabajador', 'created_by').all()
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['numero_contrato', 'trabajador__nombres', 'trabajador__apellidos', 'cargo']
    filterset_fields = ['estado', 'trabajador', 'tipo_contrato']
    ordering = ['-fecha_inicio']
    finca_field = 'trabajador__finca'  # Filtrar por finca del trabajador

    def get_serializer_class(self):
        """Seleccionar serializer según acción"""
        if self.action == 'list':
            return ContratoListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return ContratoCreateUpdateSerializer
        elif self.action in ['finalizar', 'liquidar', 'cancelar', 'reactivar']:
            # Los serializers específicos se usan en cada action
            return ContratoDetailSerializer
        return ContratoDetailSerializer

    def get_queryset(self):
        """Filtros personalizados"""
        queryset = super().get_queryset()

        # Filtro: Por vencer
        por_vencer = self.request.query_params.get('por_vencer', None)
        if por_vencer and por_vencer.lower() == 'true':
            from datetime import date, timedelta
            fecha_limite = date.today() + timedelta(days=15)
            queryset = queryset.filter(
                estado=Contrato.ACTIVO,
                fecha_fin__lte=fecha_limite,
                fecha_fin__gte=date.today()
            )

        # Filtro: Vencidos
        vencidos = self.request.query_params.get('vencidos', None)
        if vencidos and vencidos.lower() == 'true':
            from datetime import date
            queryset = queryset.filter(
                estado=Contrato.ACTIVO,
                fecha_fin__lt=date.today()
            )

        return queryset

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Estadísticas generales de contratos"""
        from datetime import date, timedelta

        total_contratos = Contrato.objects.count()
        activos = Contrato.objects.filter(estado=Contrato.ACTIVO).count()
        finalizados = Contrato.objects.filter(estado=Contrato.FINALIZADO).count()
        liquidados = Contrato.objects.filter(estado=Contrato.LIQUIDADO).count()
        cancelados = Contrato.objects.filter(estado=Contrato.CANCELADO).count()

        # Por vencer (≤15 días)
        fecha_limite = date.today() + timedelta(days=15)
        por_vencer = Contrato.objects.filter(
            estado=Contrato.ACTIVO,
            fecha_fin__lte=fecha_limite,
            fecha_fin__gte=date.today()
        ).count()

        # Vencidos
        vencidos = Contrato.objects.filter(
            estado=Contrato.ACTIVO,
            fecha_fin__lt=date.today()
        ).count()

        return Response({
            'total_contratos': total_contratos,
            'activos': activos,
            'finalizados': finalizados,
            'liquidados': liquidados,
            'cancelados': cancelados,
            'por_vencer': por_vencer,
            'vencidos': vencidos
        })

    @action(detail=False, methods=['get'])
    def alertas(self, request):
        """Contratos por vencer y vencidos"""
        from datetime import date, timedelta

        fecha_limite = date.today() + timedelta(days=15)

        # Por vencer
        por_vencer = Contrato.objects.filter(
            estado=Contrato.ACTIVO,
            fecha_fin__lte=fecha_limite,
            fecha_fin__gte=date.today()
        ).select_related('trabajador')

        # Vencidos
        vencidos = Contrato.objects.filter(
            estado=Contrato.ACTIVO,
            fecha_fin__lt=date.today()
        ).select_related('trabajador')

        por_vencer_serializer = ContratoListSerializer(por_vencer, many=True, context={'request': request})
        vencidos_serializer = ContratoListSerializer(vencidos, many=True, context={'request': request})

        return Response({
            'por_vencer': por_vencer_serializer.data,
            'vencidos': vencidos_serializer.data
        })

    @action(detail=False, methods=['get'], url_path='trabajador/(?P<trabajador_id>[^/.]+)')
    def por_trabajador(self, request, trabajador_id=None):
        """Historial de contratos de un trabajador"""
        contratos = Contrato.objects.filter(
            trabajador_id=trabajador_id
        ).select_related('trabajador', 'created_by').order_by('-fecha_inicio')

        serializer = ContratoListSerializer(contratos, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def finalizar(self, request, pk=None):
        """Finalizar contrato (ACTIVO → FINALIZADO)"""
        contrato = self.get_object()
        serializer = ContratoFinalizarSerializer(data=request.data)

        if serializer.is_valid():
            try:
                contrato.finalizar_contrato(
                    motivo=serializer.validated_data['motivo'],
                    observaciones=serializer.validated_data.get('observaciones', ''),
                    usuario=request.user
                )

                detail_serializer = ContratoDetailSerializer(contrato, context={'request': request})
                return Response(detail_serializer.data)

            except Exception as e:
                return Response(
                    {'error': str(e)},
                    status=status.HTTP_400_BAD_REQUEST
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def liquidar(self, request, pk=None):
        """Liquidar contrato (FINALIZADO → LIQUIDADO)"""
        contrato = self.get_object()
        serializer = ContratoLiquidarSerializer(data=request.data)

        if serializer.is_valid():
            try:
                contrato.liquidar_contrato(
                    fecha_liquidacion=serializer.validated_data.get('fecha_liquidacion'),
                    observaciones=serializer.validated_data.get('observaciones', ''),
                    usuario=request.user
                )

                detail_serializer = ContratoDetailSerializer(contrato, context={'request': request})
                return Response(detail_serializer.data)

            except Exception as e:
                return Response(
                    {'error': str(e)},
                    status=status.HTTP_400_BAD_REQUEST
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def cancelar(self, request, pk=None):
        """Cancelar contrato (ACTIVO → CANCELADO)"""
        contrato = self.get_object()
        serializer = ContratoCancelarSerializer(data=request.data)

        if serializer.is_valid():
            try:
                contrato.cancelar_contrato(
                    motivo=serializer.validated_data['motivo'],
                    observaciones=serializer.validated_data.get('observaciones', ''),
                    usuario=request.user
                )

                detail_serializer = ContratoDetailSerializer(contrato, context={'request': request})
                return Response(detail_serializer.data)

            except Exception as e:
                return Response(
                    {'error': str(e)},
                    status=status.HTTP_400_BAD_REQUEST
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def reactivar(self, request, pk=None):
        """Reactivar contrato (CANCELADO → ACTIVO)"""
        contrato = self.get_object()
        serializer = ContratoReactivarSerializer(data=request.data)

        if serializer.is_valid():
            try:
                contrato.reactivar_contrato(
                    observaciones=serializer.validated_data.get('observaciones', ''),
                    usuario=request.user
                )

                detail_serializer = ContratoDetailSerializer(contrato, context={'request': request})
                return Response(detail_serializer.data)

            except Exception as e:
                return Response(
                    {'error': str(e)},
                    status=status.HTTP_400_BAD_REQUEST
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'])
    def generar_pdf(self, request, pk=None):
        """Generar PDF del contrato laboral"""
        from .services.contract_pdf_generator import generar_contrato_pdf

        contrato = self.get_object()

        try:
            pdf_buffer = generar_contrato_pdf(contrato)

            response = FileResponse(
                pdf_buffer,
                content_type='application/pdf',
                as_attachment=False,
                filename=f'contrato_{contrato.numero_contrato}.pdf'
            )
            response['Content-Disposition'] = f'inline; filename="contrato_{contrato.numero_contrato}.pdf"'

            return response

        except Exception as e:
            return Response(
                {'error': f'Error al generar PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def descargar_pdf(self, request, pk=None):
        """Descargar PDF del contrato laboral"""
        from .services.contract_pdf_generator import generar_contrato_pdf

        contrato = self.get_object()

        try:
            pdf_buffer = generar_contrato_pdf(contrato)

            response = FileResponse(
                pdf_buffer,
                content_type='application/pdf',
                as_attachment=True,
                filename=f'contrato_{contrato.numero_contrato}.pdf'
            )

            return response

        except Exception as e:
            return Response(
                {'error': f'Error al generar PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def historial(self, request, pk=None):
        """Obtener historial de cambios del contrato desde auditoría"""
        contrato = self.get_object()

        # Buscar en auditoría los cambios relacionados con este contrato
        logs = AuditoriaLog.objects.filter(
            tabla_afectada='Contrato',
            registro_id=str(contrato.id)
        ).select_related('usuario').order_by('-created_at')

        historial = []
        for log in logs:
            historial.append({
                'accion': log.get_accion_display(),
                'fecha': log.created_at,
                'usuario': log.usuario.get_full_name() if log.usuario else 'Sistema',
                'descripcion': self._format_cambios(log.datos_anteriores, log.datos_nuevos)
            })

        # Agregar evento de creación si no hay logs
        if not historial:
            historial.append({
                'accion': 'Creación',
                'fecha': contrato.created_at,
                'usuario': contrato.created_by.get_full_name() if contrato.created_by else 'Sistema',
                'descripcion': f'Contrato {contrato.numero_contrato} creado'
            })

        return Response(historial)

    def _format_cambios(self, datos_ant, datos_new):
        """Formatear cambios para mostrar en historial"""
        if not datos_ant and not datos_new:
            return 'Sin detalles'

        cambios = []
        if datos_new:
            for key, value in datos_new.items():
                if key in ['created_at', 'updated_at', 'id']:
                    continue
                old_val = datos_ant.get(key, 'N/A') if datos_ant else 'N/A'
                if old_val != value:
                    cambios.append(f'{key}: {old_val} → {value}')

        return ', '.join(cambios) if cambios else 'Actualización de datos'

    @action(detail=True, methods=['get'])
    def nominas(self, request, pk=None):
        """Obtener nóminas del trabajador asociado al contrato"""
        contrato = self.get_object()
        trabajador = contrato.trabajador

        # Buscar nóminas del trabajador durante el período del contrato
        nominas_qs = Nomina.objects.filter(
            trabajador=trabajador
        ).select_related('quincena').order_by('-quincena__año', '-quincena__mes', '-quincena__numero')

        # Filtrar por fechas del contrato si aplica
        if contrato.fecha_inicio:
            nominas_qs = nominas_qs.filter(
                quincena__fecha_inicio__gte=contrato.fecha_inicio
            )
        if contrato.fecha_fin:
            nominas_qs = nominas_qs.filter(
                quincena__fecha_fin__lte=contrato.fecha_fin
            )

        # Serializar las nóminas
        nominas_data = []
        for nomina in nominas_qs[:20]:  # Limitar a 20 más recientes
            nominas_data.append({
                'id': nomina.id,
                'quincena_info': {
                    'id': nomina.quincena.id,
                    'nombre': f'Q{nomina.quincena.numero} - {nomina.quincena.mes}/{nomina.quincena.año}',
                    'fecha_inicio': nomina.quincena.fecha_inicio,
                    'fecha_fin': nomina.quincena.fecha_fin,
                },
                'total_devengado': nomina.total_devengado,
                'total_deducciones': nomina.total_deducciones,
                'total_neto': nomina.total_neto,
                'total_pagar': nomina.total_neto,  # Alias para compatibilidad
                'estado': nomina.estado,
                'estado_display': nomina.get_estado_display(),
                'fecha_calculo': nomina.fecha_calculo,
            })

        return Response(nominas_data)


class DocumentoContratoViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de documentos de contratos"""
    queryset = DocumentoContrato.objects.select_related('contrato', 'created_by').all()
    serializer_class = DocumentoContratoSerializer
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['contrato', 'tipo_documento']
    ordering = ['-created_at']

    def perform_create(self, serializer):
        """Guardar usuario que sube el documento"""
        serializer.save(created_by=self.request.user)


# ============================================================================
# CONFIGURACIÓN EMPRESA
# ============================================================================

class ConfiguracionEmpresaViewSet(viewsets.ModelViewSet):
    """
    ViewSet para la configuración de empresa (singleton).
    Solo puede existir un registro.
    """
    queryset = ConfiguracionEmpresa.objects.all()
    serializer_class = ConfiguracionEmpresaSerializer
    permission_classes = [CanModifyData]

    def list(self, request, *args, **kwargs):
        """Retorna la configuración única de empresa"""
        config = ConfiguracionEmpresa.get_config()
        serializer = self.get_serializer(config)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        """Crear o actualizar la configuración única"""
        config = ConfiguracionEmpresa.get_config()
        serializer = self.get_serializer(config, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def actual(self, request):
        """Obtener la configuración actual de empresa"""
        config = ConfiguracionEmpresa.get_config()
        serializer = self.get_serializer(config)
        return Response(serializer.data)


# ============================================================================
# PILA - SEGURIDAD SOCIAL
# ============================================================================

class PILAViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de PILA (Planilla Integrada de Liquidación de Aportes).
    Calcula aportes de seguridad social basados en las nóminas del mes.
    """
    queryset = PILA.objects.all()
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['año', 'mes', 'estado', 'tipo']
    ordering = ['-año', '-mes']

    def get_serializer_class(self):
        if self.action == 'list':
            return PILAListSerializer
        elif self.action == 'create':
            return PILACreateSerializer
        return PILADetailSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['post'])
    def calcular(self, request):
        """
        Calcular PILA basado en las nóminas aprobadas.
        Parámetros:
          - tipo: 'MES' (default) o 'QUINCENA'
          - mes, año: requeridos siempre
          - quincena_id: requerido si tipo='QUINCENA'
          - fincas: lista opcional de IDs de finca para filtrar
        """
        from django.db.models import Sum
        from decimal import Decimal
        from collections import defaultdict

        mes = request.data.get('mes')
        año = request.data.get('año')
        tipo = request.data.get('tipo', 'MES')
        quincena_id = request.data.get('quincena_id')
        fincas_ids = request.data.get('fincas', [])

        if not mes or not año:
            return Response(
                {'error': 'Se requieren mes y año'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar fincas contra permisos del usuario
        if fincas_ids:
            user = request.user
            if not user.es_administrador:
                fincas_permitidas = set(user.fincas_asignadas.values_list('id', flat=True))
                fincas_solicitadas = set(int(f) for f in fincas_ids)
                if not fincas_solicitadas.issubset(fincas_permitidas):
                    return Response(
                        {'error': 'No tiene permisos para algunas fincas seleccionadas'},
                        status=status.HTTP_403_FORBIDDEN
                    )

        if tipo == 'QUINCENA':
            # Calcular por quincena individual
            if not quincena_id:
                return Response(
                    {'error': 'Se requiere quincena_id para cálculo quincenal'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            try:
                quincena_obj = Quincena.objects.get(id=quincena_id)
            except Quincena.DoesNotExist:
                return Response(
                    {'error': 'Quincena no encontrada'},
                    status=status.HTTP_404_NOT_FOUND
                )

            if PILA.objects.filter(quincena=quincena_obj, tipo='QUINCENA').exists():
                return Response(
                    {'error': f'Ya existe una PILA para esa quincena'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            quincenas_list = [quincena_obj]
            dias_por_quincena = 15
        else:
            # Calcular por mes completo
            if PILA.objects.filter(mes=mes, año=año, tipo='MES').exists():
                return Response(
                    {'error': f'Ya existe una PILA mensual para {mes:02d}/{año}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            quincenas_list = list(Quincena.objects.filter(mes=mes, año=año))
            if not quincenas_list:
                return Response(
                    {'error': f'No hay quincenas para {mes:02d}/{año}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            dias_por_quincena = 15

        # Obtener nóminas aprobadas
        nominas = Nomina.objects.filter(
            quincena__in=quincenas_list,
            estado='APROBADA'
        ).select_related('trabajador', 'quincena')

        # Filtrar por fincas si se especificaron
        if fincas_ids:
            nominas = nominas.filter(trabajador__finca__id__in=fincas_ids)

        if not nominas.exists():
            return Response(
                {'error': 'No hay nóminas aprobadas para el período seleccionado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Crear PILA
        pila = PILA.objects.create(
            mes=mes,
            año=año,
            tipo=tipo,
            quincena=quincenas_list[0] if tipo == 'QUINCENA' else None,
            estado='CALCULADA',
            created_by=request.user
        )

        # Agrupar nóminas por trabajador y calcular IBC
        trabajadores_ibc = defaultdict(lambda: {'ibc': Decimal('0'), 'dias': 0, 'nominas': []})

        for nomina in nominas:
            trab_id = nomina.trabajador_id
            auxilio = DetalleNomina.objects.filter(
                nomina=nomina,
                concepto='AUXILIO_TRANSPORTE'
            ).aggregate(total=Sum('valor_total'))['total'] or Decimal('0')
            ibc = nomina.total_devengado - auxilio
            trabajadores_ibc[trab_id]['ibc'] += ibc
            trabajadores_ibc[trab_id]['dias'] += dias_por_quincena
            trabajadores_ibc[trab_id]['nominas'].append(nomina)

        # Crear detalles por trabajador
        total_ibc = Decimal('0')
        total_salud = Decimal('0')
        total_pension = Decimal('0')
        total_arl = Decimal('0')
        total_caja = Decimal('0')

        max_dias = 15 if tipo == 'QUINCENA' else 30

        for trab_id, data in trabajadores_ibc.items():
            ibc = data['ibc']
            dias = min(data['dias'], max_dias)

            aporte_salud = ibc * PORCENTAJE_SALUD_EMPLEADOR
            aporte_pension = ibc * PORCENTAJE_PENSION_EMPLEADOR
            aporte_arl = ibc * PORCENTAJE_ARL
            aporte_caja = ibc * PORCENTAJE_CAJA_COMPENSACION
            total_trabajador = aporte_salud + aporte_pension + aporte_arl + aporte_caja

            DetallePILA.objects.create(
                pila=pila,
                trabajador_id=trab_id,
                ibc=ibc,
                dias_cotizados=dias,
                aporte_salud=aporte_salud,
                aporte_pension=aporte_pension,
                aporte_arl=aporte_arl,
                aporte_caja=aporte_caja,
                total_aportes=total_trabajador,
                nomina=data['nominas'][-1]
            )

            total_ibc += ibc
            total_salud += aporte_salud
            total_pension += aporte_pension
            total_arl += aporte_arl
            total_caja += aporte_caja

        # Actualizar totales en PILA
        pila.total_ibc = total_ibc
        pila.total_salud = total_salud
        pila.total_pension = total_pension
        pila.total_arl = total_arl
        pila.total_caja = total_caja
        pila.total_aportes = total_salud + total_pension + total_arl + total_caja
        pila.save()

        serializer = PILADetailSerializer(pila)
        return Response({
            'message': f'PILA calculada con {len(trabajadores_ibc)} trabajadores',
            'pila': serializer.data
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def marcar_pagada(self, request, pk=None):
        """Marcar PILA como pagada"""
        pila = self.get_object()

        if pila.estado == 'PAGADA':
            return Response(
                {'error': 'Esta PILA ya está marcada como pagada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = PILAMarcarPagadaSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        pila.estado = 'PAGADA'
        pila.fecha_pago = serializer.validated_data['fecha_pago']
        pila.numero_planilla = serializer.validated_data['numero_planilla']
        pila.save()

        return Response({
            'message': 'PILA marcada como pagada',
            'pila': PILADetailSerializer(pila).data
        })

    @action(detail=False, methods=['get'])
    def resumen(self, request):
        """Obtener resumen de PILA pendientes y porcentajes vigentes"""
        from datetime import date
        hoy = date.today()

        # PILA del mes actual
        pila_actual = PILA.objects.filter(mes=hoy.month, año=hoy.year).first()

        # PILAs pendientes (no pagadas)
        pilas_pendientes = PILA.objects.filter(estado__in=['BORRADOR', 'CALCULADA'])
        total_pendiente = sum(p.total_aportes for p in pilas_pendientes)

        # Porcentajes vigentes
        porcentajes = {
            'seguridad_social': {
                'salud': float(PORCENTAJE_SALUD_EMPLEADOR * 100),
                'pension': float(PORCENTAJE_PENSION_EMPLEADOR * 100),
                'arl': float(PORCENTAJE_ARL * 100),
                'caja_compensacion': float(PORCENTAJE_CAJA_COMPENSACION * 100),
                'total': float((PORCENTAJE_SALUD_EMPLEADOR + PORCENTAJE_PENSION_EMPLEADOR +
                               PORCENTAJE_ARL + PORCENTAJE_CAJA_COMPENSACION) * 100)
            },
            'prestaciones': {
                'cesantias': float(PORCENTAJE_CESANTIAS * 100),
                'intereses_cesantias': float(PORCENTAJE_INTERESES_CESANTIAS * 100),
                'prima': float(PORCENTAJE_PRIMA * 100),
                'vacaciones': float(PORCENTAJE_VACACIONES * 100),
                'total': float((PORCENTAJE_CESANTIAS + PORCENTAJE_INTERESES_CESANTIAS +
                               PORCENTAJE_PRIMA + PORCENTAJE_VACACIONES) * 100)
            }
        }

        return Response({
            'pila_mes_actual': PILAListSerializer(pila_actual).data if pila_actual else None,
            'pilas_pendientes': pilas_pendientes.count(),
            'total_pendiente': total_pendiente,
            'porcentajes': porcentajes
        })

    @action(detail=True, methods=['get'])
    def exportar_excel(self, request, pk=None):
        """Exportar PILA a Excel. Acepta ?fincas=1,2,3 para filtrar por fincas."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from decimal import Decimal

        pila = self.get_object()
        detalles = pila.detalles.select_related('trabajador').order_by(
            'trabajador__apellidos', 'trabajador__nombres'
        )

        # Filtrar por fincas si se especificaron
        fincas_param = request.query_params.get('fincas', '')
        if fincas_param:
            fincas_ids = [int(f) for f in fincas_param.split(',') if f.strip()]
            detalles = detalles.filter(trabajador__finca__id__in=fincas_ids)

        wb = Workbook()
        ws = wb.active
        ws.title = "PILA"

        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="1F4E78")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center = Alignment(horizontal='center', vertical='center')
        right = Alignment(horizontal='right', vertical='center')

        meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

        # Título
        ws.merge_cells('A1:I1')
        cell = ws['A1']
        cell.value = f"PILA - {pila.periodo_display}"
        cell.font = title_font
        cell.alignment = center

        # Encabezados
        headers = ['Trabajador', 'Documento', 'IBC', 'Días', 'Salud (8.5%)',
                   'Pensión (12%)', 'ARL (1.044%)', 'Caja (4%)', 'Total Aportes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # Datos y totales calculados sobre detalles filtrados
        row = 4
        sum_ibc = Decimal('0')
        sum_salud = Decimal('0')
        sum_pension = Decimal('0')
        sum_arl = Decimal('0')
        sum_caja = Decimal('0')
        sum_total = Decimal('0')

        for detalle in detalles:
            t = detalle.trabajador
            ws.cell(row=row, column=1, value=t.nombre_completo).border = border
            ws.cell(row=row, column=2, value=t.numero_documento).border = border

            for col, val in [(3, detalle.ibc), (5, detalle.aporte_salud),
                             (6, detalle.aporte_pension), (7, detalle.aporte_arl),
                             (8, detalle.aporte_caja), (9, detalle.total_aportes)]:
                c = ws.cell(row=row, column=col, value=float(val))
                c.number_format = '$#,##0'
                c.alignment = right
                c.border = border

            c = ws.cell(row=row, column=4, value=detalle.dias_cotizados)
            c.alignment = center
            c.border = border

            sum_ibc += detalle.ibc
            sum_salud += detalle.aporte_salud
            sum_pension += detalle.aporte_pension
            sum_arl += detalle.aporte_arl
            sum_caja += detalle.aporte_caja
            sum_total += detalle.total_aportes

            row += 1

        # Totales
        total_font = Font(bold=True)
        total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        ws.cell(row=row, column=1, value="TOTALES").font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=1).border = border

        for col, val in [(3, sum_ibc), (5, sum_salud),
                         (6, sum_pension), (7, sum_arl),
                         (8, sum_caja), (9, sum_total)]:
            c = ws.cell(row=row, column=col, value=float(val))
            c.number_format = '$#,##0'
            c.font = total_font
            c.fill = total_fill
            c.alignment = right
            c.border = border

        for col in [2, 4]:
            ws.cell(row=row, column=col).fill = total_fill
            ws.cell(row=row, column=col).border = border

        # Anchos
        anchos = {'A': 35, 'B': 18, 'C': 18, 'D': 8, 'E': 16,
                  'F': 16, 'G': 16, 'H': 16, 'I': 18}
        for col_letter, ancho in anchos.items():
            ws.column_dimensions[col_letter].width = ancho

        # Generar response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"PILA_{meses[pila.mes]}_{pila.año}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ============================================================================
# PRESTACIONES SOCIALES
# ============================================================================

class PrestacionesViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de provisiones de prestaciones sociales.
    Calcula provisiones mensuales basadas en las nóminas.
    """
    queryset = ResumenPrestaciones.objects.all()
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['año', 'mes']
    ordering = ['-año', '-mes']

    def get_serializer_class(self):
        if self.action == 'list':
            return ResumenPrestacionesListSerializer
        elif self.action == 'create':
            return ResumenPrestacionesCreateSerializer
        return ResumenPrestacionesDetailSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['post'])
    def calcular(self, request):
        """
        Calcular provisiones de prestaciones para un mes específico.
        Requiere: mes, año
        """
        mes = request.data.get('mes')
        año = request.data.get('año')

        if not mes or not año:
            return Response(
                {'error': 'Se requieren mes y año'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar si ya existe
        if ResumenPrestaciones.objects.filter(mes=mes, año=año).exists():
            return Response(
                {'error': f'Ya existen provisiones para {mes:02d}/{año}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener quincenas del mes
        from decimal import Decimal
        quincenas = Quincena.objects.filter(mes=mes, año=año)

        if not quincenas.exists():
            return Response(
                {'error': f'No hay quincenas para {mes:02d}/{año}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener nóminas aprobadas
        nominas = Nomina.objects.filter(
            quincena__in=quincenas,
            estado='APROBADA'
        ).select_related('trabajador', 'quincena')

        if not nominas.exists():
            return Response(
                {'error': f'No hay nóminas aprobadas para {mes:02d}/{año}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Agrupar por trabajador
        from collections import defaultdict
        trabajadores_salario = defaultdict(lambda: {'salario': Decimal('0'), 'nominas': []})

        for nomina in nominas:
            trab_id = nomina.trabajador_id
            trabajadores_salario[trab_id]['salario'] += nomina.total_devengado
            trabajadores_salario[trab_id]['nominas'].append(nomina)

        # Calcular provisiones por trabajador
        total_salario = Decimal('0')
        total_cesantias = Decimal('0')
        total_intereses = Decimal('0')
        total_prima = Decimal('0')
        total_vacaciones = Decimal('0')

        provisiones_creadas = []

        for trab_id, data in trabajadores_salario.items():
            salario = data['salario']

            # Calcular provisiones
            cesantias = salario * PORCENTAJE_CESANTIAS
            intereses = cesantias * PORCENTAJE_INTERESES_CESANTIAS
            prima = salario * PORCENTAJE_PRIMA
            vacaciones = salario * PORCENTAJE_VACACIONES
            total_provision = cesantias + intereses + prima + vacaciones

            # Crear provisión
            provision = ProvisionPrestaciones.objects.create(
                mes=mes,
                año=año,
                trabajador_id=trab_id,
                salario_base=salario,
                cesantias=cesantias,
                intereses_cesantias=intereses,
                prima=prima,
                vacaciones=vacaciones,
                total_provision=total_provision,
                nomina=data['nominas'][-1]
            )
            provisiones_creadas.append(provision)

            # Acumular totales
            total_salario += salario
            total_cesantias += cesantias
            total_intereses += intereses
            total_prima += prima
            total_vacaciones += vacaciones

        # Crear resumen
        resumen = ResumenPrestaciones.objects.create(
            mes=mes,
            año=año,
            total_salario_base=total_salario,
            total_cesantias=total_cesantias,
            total_intereses_cesantias=total_intereses,
            total_prima=total_prima,
            total_vacaciones=total_vacaciones,
            total_provisiones=total_cesantias + total_intereses + total_prima + total_vacaciones,
            num_trabajadores=len(trabajadores_salario),
            created_by=request.user
        )

        serializer = ResumenPrestacionesDetailSerializer(resumen)
        return Response({
            'message': f'Prestaciones calculadas para {mes:02d}/{año} con {len(trabajadores_salario)} trabajadores',
            'resumen': serializer.data
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def acumulado_año(self, request):
        """Obtener acumulado de provisiones del año"""
        from datetime import date
        from django.db.models import Sum

        año = request.query_params.get('año', date.today().year)

        acumulado = ResumenPrestaciones.objects.filter(año=año).aggregate(
            total_cesantias=Sum('total_cesantias'),
            total_intereses=Sum('total_intereses_cesantias'),
            total_prima=Sum('total_prima'),
            total_vacaciones=Sum('total_vacaciones'),
            total_provisiones=Sum('total_provisiones'),
            total_salarios=Sum('total_salario_base')
        )

        meses_calculados = ResumenPrestaciones.objects.filter(año=año).count()

        return Response({
            'año': año,
            'meses_calculados': meses_calculados,
            'acumulado': acumulado
        })

    @action(detail=False, methods=['get'])
    def por_trabajador(self, request):
        """Obtener provisiones acumuladas por trabajador"""
        from django.db.models import Sum
        from datetime import date

        año = request.query_params.get('año', date.today().year)
        trabajador_id = request.query_params.get('trabajador')

        queryset = ProvisionPrestaciones.objects.filter(año=año)

        if trabajador_id:
            queryset = queryset.filter(trabajador_id=trabajador_id)

        acumulado = queryset.values(
            'trabajador__id',
            'trabajador__nombres',
            'trabajador__apellidos',
            'trabajador__numero_documento'
        ).annotate(
            total_cesantias=Sum('cesantias'),
            total_intereses=Sum('intereses_cesantias'),
            total_prima=Sum('prima'),
            total_vacaciones=Sum('vacaciones'),
            total_provisiones=Sum('total_provision'),
            total_salarios=Sum('salario_base')
        ).order_by('trabajador__apellidos', 'trabajador__nombres')

        return Response({
            'año': año,
            'trabajadores': list(acumulado)
        })
