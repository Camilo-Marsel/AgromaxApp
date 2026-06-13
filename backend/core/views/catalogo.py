# backend/core/views/catalogo.py

from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from django.utils import timezone

from ..models import UnidadMedida, Labor, LaborInsumo, ListaPrecios
from ..serializers import (
    UnidadMedidaSerializer, LaborListSerializer, LaborCreateUpdateSerializer,
    LaborInsumoSerializer, LaborInsumoCreateUpdateSerializer,
    ListaPreciosSerializer, ListaPreciosCreateSerializer,
)
from ..permissions import CanModifyData
from ..filters import LaborFilter


class UnidadMedidaViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet de solo lectura para Unidades de Medida"""
    queryset = UnidadMedida.objects.all()
    serializer_class = UnidadMedidaSerializer
    permission_classes = [permissions.IsAuthenticated]


class LaborViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de labores"""
    queryset = Labor.objects.all()
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = LaborFilter
    search_fields = ['codigo', 'nombre']
    ordering_fields = ['codigo', 'nombre', 'created_at']
    ordering = ['nombre']

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return LaborCreateUpdateSerializer
        return LaborListSerializer

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """Exportar listado de labores a Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        queryset = self.filter_queryset(self.get_queryset())
        labores = list(queryset.select_related('unidad_medida'))

        wb = Workbook()
        ws = wb.active
        ws.title = "Listado de Labores"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="1F4E78")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')

        ws.merge_cells('A1:E1')
        cell = ws['A1']
        cell.value = "LISTADO DE LABORES Y PRECIOS"
        cell.font = title_font
        cell.alignment = center_alignment
        ws.append([])

        headers = ['Código', 'Nombre', 'Unidad de Medida', 'Precio Actual', 'Estado']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        row = 4
        hoy = timezone.now().date()
        for labor in labores:
            ws.cell(row=row, column=1, value=labor.codigo)
            ws.cell(row=row, column=2, value=labor.nombre)
            ws.cell(row=row, column=3, value=labor.unidad_medida.get_nombre_display())

            precio = labor.precios.filter(
                fecha_inicio_vigencia__lte=hoy,
                fecha_fin_vigencia__isnull=True
            ).first()

            cell_precio = ws.cell(row=row, column=4)
            if precio:
                cell_precio.value = float(precio.precio)
                cell_precio.number_format = '$#,##0.00'
            else:
                cell_precio.value = 'Sin precio'
            cell_precio.alignment = right_alignment

            ws.cell(row=row, column=5, value='Activa' if labor.activa else 'Inactiva')

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border

            row += 1

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        fecha_str = timezone.now().strftime('%Y%m%d')
        filename = f"Listado_Labores_{fecha_str}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ListaPreciosViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de precios"""
    queryset = ListaPrecios.objects.all()
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['labor', 'fecha_inicio_vigencia']
    ordering_fields = ['fecha_inicio_vigencia', 'created_at']
    ordering = ['-fecha_inicio_vigencia']

    def get_serializer_class(self):
        if self.action == 'create':
            return ListaPreciosCreateSerializer
        return ListaPreciosSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class LaborInsumoViewSet(viewsets.ModelViewSet):
    queryset = LaborInsumo.objects.select_related('labor', 'producto').all()
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['labor']
    ordering_fields = ['labor__nombre', 'created_at']
    ordering = ['labor__nombre']

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return LaborInsumoCreateUpdateSerializer
        return LaborInsumoSerializer
