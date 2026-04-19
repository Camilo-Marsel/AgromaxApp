# backend/core/views/obligaciones.py

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.filters import OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from decimal import Decimal
from collections import defaultdict

from ..models import (
    PILA, DetallePILA, ProvisionPrestaciones, ResumenPrestaciones,
    Nomina, Quincena, DetalleNomina,
    PORCENTAJE_SALUD_EMPLEADOR, PORCENTAJE_PENSION_EMPLEADOR,
    PORCENTAJE_ARL, PORCENTAJE_CAJA_COMPENSACION,
    PORCENTAJE_CESANTIAS, PORCENTAJE_INTERESES_CESANTIAS,
    PORCENTAJE_PRIMA, PORCENTAJE_VACACIONES,
)
from ..serializers import (
    PILAListSerializer, PILADetailSerializer, PILACreateSerializer,
    PILAMarcarPagadaSerializer, ResumenPrestacionesListSerializer,
    ResumenPrestacionesDetailSerializer, ResumenPrestacionesCreateSerializer,
)
from ..permissions import CanModifyData


class PILAViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de PILA (Planilla Integrada de Liquidación de Aportes).
    Calcula aportes de seguridad social basados en las nóminas del mes.
    """
    queryset = PILA.objects.all()
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['año', 'mes', 'estado', 'tipo']
    ordering = ['-año', '-mes']

    def get_serializer_class(self):
        if self.action == 'list':
            return PILAListSerializer
        elif self.action == 'create':
            return PILACreateSerializer
        return PILADetailSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['post'])
    def calcular(self, request):
        """Calcular PILA basado en las nóminas aprobadas.

        Parámetros:
          - tipo: 'MES' (default) o 'QUINCENA'
          - mes, año: requeridos siempre
          - quincena_id: requerido si tipo='QUINCENA'
          - fincas: lista opcional de IDs de finca para filtrar
        """
        from django.db.models import Sum

        mes = request.data.get('mes')
        año = request.data.get('año')
        tipo = request.data.get('tipo', 'MES')
        quincena_id = request.data.get('quincena_id')
        fincas_ids = request.data.get('fincas', [])

        if not mes or not año:
            return Response(
                {'error': 'Se requieren mes y año'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if fincas_ids:
            user = request.user
            if not user.es_administrador:
                fincas_permitidas = set(user.fincas_asignadas.values_list('id', flat=True))
                fincas_solicitadas = set(int(f) for f in fincas_ids)
                if not fincas_solicitadas.issubset(fincas_permitidas):
                    return Response(
                        {'error': 'No tiene permisos para algunas fincas seleccionadas'},
                        status=status.HTTP_403_FORBIDDEN
                    )

        if tipo == 'QUINCENA':
            if not quincena_id:
                return Response(
                    {'error': 'Se requiere quincena_id para cálculo quincenal'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            try:
                quincena_obj = Quincena.objects.get(id=quincena_id)
            except Quincena.DoesNotExist:
                return Response(
                    {'error': 'Quincena no encontrada'},
                    status=status.HTTP_404_NOT_FOUND
                )

            if PILA.objects.filter(quincena=quincena_obj, tipo='QUINCENA').exists():
                return Response(
                    {'error': 'Ya existe una PILA para esa quincena'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            quincenas_list = [quincena_obj]
            dias_por_quincena = 15
        else:
            if PILA.objects.filter(mes=mes, año=año, tipo='MES').exists():
                return Response(
                    {'error': f'Ya existe una PILA mensual para {mes:02d}/{año}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            quincenas_list = list(Quincena.objects.filter(mes=mes, año=año))
            if not quincenas_list:
                return Response(
                    {'error': f'No hay quincenas para {mes:02d}/{año}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            dias_por_quincena = 15

        nominas = Nomina.objects.filter(
            quincena__in=quincenas_list,
            estado='APROBADA'
        ).select_related('trabajador', 'quincena')

        if fincas_ids:
            nominas = nominas.filter(trabajador__finca__id__in=fincas_ids)

        if not nominas.exists():
            return Response(
                {'error': 'No hay nóminas aprobadas para el período seleccionado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        pila = PILA.objects.create(
            mes=mes,
            año=año,
            tipo=tipo,
            quincena=quincenas_list[0] if tipo == 'QUINCENA' else None,
            estado='CALCULADA',
            created_by=request.user
        )

        trabajadores_ibc = defaultdict(lambda: {'ibc': Decimal('0'), 'dias': 0, 'nominas': []})

        for nomina in nominas:
            trab_id = nomina.trabajador_id
            auxilio = DetalleNomina.objects.filter(
                nomina=nomina,
                concepto='AUXILIO_TRANSPORTE'
            ).aggregate(total=Sum('valor_total'))['total'] or Decimal('0')
            ibc = nomina.total_devengado - auxilio
            trabajadores_ibc[trab_id]['ibc'] += ibc
            trabajadores_ibc[trab_id]['dias'] += dias_por_quincena
            trabajadores_ibc[trab_id]['nominas'].append(nomina)

        total_ibc = Decimal('0')
        total_salud = Decimal('0')
        total_pension = Decimal('0')
        total_arl = Decimal('0')
        total_caja = Decimal('0')

        max_dias = 15 if tipo == 'QUINCENA' else 30

        for trab_id, data in trabajadores_ibc.items():
            ibc = data['ibc']
            if ibc <= 0:
                continue
            dias = min(data['dias'], max_dias)

            aporte_salud = ibc * PORCENTAJE_SALUD_EMPLEADOR
            aporte_pension = ibc * PORCENTAJE_PENSION_EMPLEADOR
            aporte_arl = ibc * PORCENTAJE_ARL
            aporte_caja = ibc * PORCENTAJE_CAJA_COMPENSACION
            total_trabajador = aporte_salud + aporte_pension + aporte_arl + aporte_caja

            DetallePILA.objects.create(
                pila=pila,
                trabajador_id=trab_id,
                ibc=ibc,
                dias_cotizados=dias,
                aporte_salud=aporte_salud,
                aporte_pension=aporte_pension,
                aporte_arl=aporte_arl,
                aporte_caja=aporte_caja,
                total_aportes=total_trabajador,
                nomina=data['nominas'][-1]
            )

            total_ibc += ibc
            total_salud += aporte_salud
            total_pension += aporte_pension
            total_arl += aporte_arl
            total_caja += aporte_caja

        pila.total_ibc = total_ibc
        pila.total_salud = total_salud
        pila.total_pension = total_pension
        pila.total_arl = total_arl
        pila.total_caja = total_caja
        pila.total_aportes = total_salud + total_pension + total_arl + total_caja
        pila.save()

        serializer = PILADetailSerializer(pila)
        return Response({
            'message': f'PILA calculada con {len(trabajadores_ibc)} trabajadores',
            'pila': serializer.data
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def marcar_pagada(self, request, pk=None):
        """Marcar PILA como pagada"""
        pila = self.get_object()

        if pila.estado == 'PAGADA':
            return Response(
                {'error': 'Esta PILA ya está marcada como pagada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = PILAMarcarPagadaSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        pila.estado = 'PAGADA'
        pila.fecha_pago = serializer.validated_data['fecha_pago']
        pila.numero_planilla = serializer.validated_data['numero_planilla']
        pila.save()

        return Response({
            'message': 'PILA marcada como pagada',
            'pila': PILADetailSerializer(pila).data
        })

    @action(detail=False, methods=['get'])
    def resumen(self, request):
        """Obtener resumen de PILA pendientes y porcentajes vigentes"""
        from datetime import date
        hoy = date.today()

        pila_actual = PILA.objects.filter(mes=hoy.month, año=hoy.year).first()
        pilas_pendientes = PILA.objects.filter(estado__in=['BORRADOR', 'CALCULADA'])
        total_pendiente = sum(p.total_aportes for p in pilas_pendientes)

        porcentajes = {
            'seguridad_social': {
                'salud': float(PORCENTAJE_SALUD_EMPLEADOR * 100),
                'pension': float(PORCENTAJE_PENSION_EMPLEADOR * 100),
                'arl': float(PORCENTAJE_ARL * 100),
                'caja_compensacion': float(PORCENTAJE_CAJA_COMPENSACION * 100),
                'total': float((PORCENTAJE_SALUD_EMPLEADOR + PORCENTAJE_PENSION_EMPLEADOR +
                               PORCENTAJE_ARL + PORCENTAJE_CAJA_COMPENSACION) * 100)
            },
            'prestaciones': {
                'cesantias': float(PORCENTAJE_CESANTIAS * 100),
                'intereses_cesantias': float(PORCENTAJE_INTERESES_CESANTIAS * 100),
                'prima': float(PORCENTAJE_PRIMA * 100),
                'vacaciones': float(PORCENTAJE_VACACIONES * 100),
                'total': float((PORCENTAJE_CESANTIAS + PORCENTAJE_INTERESES_CESANTIAS +
                               PORCENTAJE_PRIMA + PORCENTAJE_VACACIONES) * 100)
            }
        }

        return Response({
            'pila_mes_actual': PILAListSerializer(pila_actual).data if pila_actual else None,
            'pilas_pendientes': pilas_pendientes.count(),
            'total_pendiente': total_pendiente,
            'porcentajes': porcentajes
        })

    @action(detail=True, methods=['get'])
    def exportar_excel(self, request, pk=None):
        """Exportar PILA a Excel. Acepta ?fincas=1,2,3 para filtrar por fincas."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        pila = self.get_object()
        detalles = pila.detalles.select_related('trabajador').order_by(
            'trabajador__apellidos', 'trabajador__nombres'
        )

        fincas_param = request.query_params.get('fincas', '')
        if fincas_param:
            fincas_ids = [int(f) for f in fincas_param.split(',') if f.strip()]
            detalles = detalles.filter(trabajador__finca__id__in=fincas_ids)

        wb = Workbook()
        ws = wb.active
        ws.title = "PILA"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="1F4E78")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center = Alignment(horizontal='center', vertical='center')
        right = Alignment(horizontal='right', vertical='center')

        meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

        ws.merge_cells('A1:I1')
        cell = ws['A1']
        cell.value = f"PILA - {pila.periodo_display}"
        cell.font = title_font
        cell.alignment = center

        headers = ['Trabajador', 'Documento', 'IBC', 'Días', 'Salud (8.5%)',
                   'Pensión (12%)', 'ARL (1.044%)', 'Caja (4%)', 'Total Aportes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        row = 4
        sum_ibc = Decimal('0')
        sum_salud = Decimal('0')
        sum_pension = Decimal('0')
        sum_arl = Decimal('0')
        sum_caja = Decimal('0')
        sum_total = Decimal('0')

        for detalle in detalles:
            t = detalle.trabajador
            ws.cell(row=row, column=1, value=t.nombre_completo).border = border
            ws.cell(row=row, column=2, value=t.numero_documento).border = border

            for col, val in [(3, detalle.ibc), (5, detalle.aporte_salud),
                             (6, detalle.aporte_pension), (7, detalle.aporte_arl),
                             (8, detalle.aporte_caja), (9, detalle.total_aportes)]:
                c = ws.cell(row=row, column=col, value=float(val))
                c.number_format = '$#,##0'
                c.alignment = right
                c.border = border

            c = ws.cell(row=row, column=4, value=detalle.dias_cotizados)
            c.alignment = center
            c.border = border

            sum_ibc += detalle.ibc
            sum_salud += detalle.aporte_salud
            sum_pension += detalle.aporte_pension
            sum_arl += detalle.aporte_arl
            sum_caja += detalle.aporte_caja
            sum_total += detalle.total_aportes

            row += 1

        total_font = Font(bold=True)
        total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        ws.cell(row=row, column=1, value="TOTALES").font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=1).border = border

        for col, val in [(3, sum_ibc), (5, sum_salud),
                         (6, sum_pension), (7, sum_arl),
                         (8, sum_caja), (9, sum_total)]:
            c = ws.cell(row=row, column=col, value=float(val))
            c.number_format = '$#,##0'
            c.font = total_font
            c.fill = total_fill
            c.alignment = right
            c.border = border

        for col in [2, 4]:
            ws.cell(row=row, column=col).fill = total_fill
            ws.cell(row=row, column=col).border = border

        anchos = {'A': 35, 'B': 18, 'C': 18, 'D': 8, 'E': 16,
                  'F': 16, 'G': 16, 'H': 16, 'I': 18}
        for col_letter, ancho in anchos.items():
            ws.column_dimensions[col_letter].width = ancho

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"PILA_{meses[pila.mes]}_{pila.año}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class PrestacionesViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de provisiones de prestaciones sociales.
    Calcula provisiones mensuales basadas en las nóminas.
    """
    queryset = ResumenPrestaciones.objects.all()
    permission_classes = [CanModifyData]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['año', 'mes']
    ordering = ['-año', '-mes']

    def get_serializer_class(self):
        if self.action == 'list':
            return ResumenPrestacionesListSerializer
        elif self.action == 'create':
            return ResumenPrestacionesCreateSerializer
        return ResumenPrestacionesDetailSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['post'])
    def calcular(self, request):
        """Calcular provisiones de prestaciones para un mes específico. Requiere: mes, año"""
        mes = request.data.get('mes')
        año = request.data.get('año')

        if not mes or not año:
            return Response(
                {'error': 'Se requieren mes y año'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if ResumenPrestaciones.objects.filter(mes=mes, año=año).exists():
            return Response(
                {'error': f'Ya existen provisiones para {mes:02d}/{año}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        quincenas = Quincena.objects.filter(mes=mes, año=año)

        if not quincenas.exists():
            return Response(
                {'error': f'No hay quincenas para {mes:02d}/{año}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        nominas = Nomina.objects.filter(
            quincena__in=quincenas,
            estado='APROBADA'
        ).select_related('trabajador', 'quincena')

        if not nominas.exists():
            return Response(
                {'error': f'No hay nóminas aprobadas para {mes:02d}/{año}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        trabajadores_salario = defaultdict(lambda: {'salario': Decimal('0'), 'nominas': []})

        for nomina in nominas:
            trab_id = nomina.trabajador_id
            trabajadores_salario[trab_id]['salario'] += nomina.total_devengado
            trabajadores_salario[trab_id]['nominas'].append(nomina)

        total_salario = Decimal('0')
        total_cesantias = Decimal('0')
        total_intereses = Decimal('0')
        total_prima = Decimal('0')
        total_vacaciones = Decimal('0')

        provisiones_creadas = []

        for trab_id, data in trabajadores_salario.items():
            salario = data['salario']

            cesantias = salario * PORCENTAJE_CESANTIAS
            intereses = cesantias * PORCENTAJE_INTERESES_CESANTIAS
            prima = salario * PORCENTAJE_PRIMA
            vacaciones = salario * PORCENTAJE_VACACIONES
            total_provision = cesantias + intereses + prima + vacaciones

            provision = ProvisionPrestaciones.objects.create(
                mes=mes,
                año=año,
                trabajador_id=trab_id,
                salario_base=salario,
                cesantias=cesantias,
                intereses_cesantias=intereses,
                prima=prima,
                vacaciones=vacaciones,
                total_provision=total_provision,
                nomina=data['nominas'][-1]
            )
            provisiones_creadas.append(provision)

            total_salario += salario
            total_cesantias += cesantias
            total_intereses += intereses
            total_prima += prima
            total_vacaciones += vacaciones

        resumen = ResumenPrestaciones.objects.create(
            mes=mes,
            año=año,
            total_salario_base=total_salario,
            total_cesantias=total_cesantias,
            total_intereses_cesantias=total_intereses,
            total_prima=total_prima,
            total_vacaciones=total_vacaciones,
            total_provisiones=total_cesantias + total_intereses + total_prima + total_vacaciones,
            num_trabajadores=len(trabajadores_salario),
            created_by=request.user
        )

        serializer = ResumenPrestacionesDetailSerializer(resumen)
        return Response({
            'message': f'Prestaciones calculadas para {mes:02d}/{año} con {len(trabajadores_salario)} trabajadores',
            'resumen': serializer.data
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def acumulado_año(self, request):
        """Obtener acumulado de provisiones del año"""
        from datetime import date
        from django.db.models import Sum

        año = request.query_params.get('año', date.today().year)

        acumulado = ResumenPrestaciones.objects.filter(año=año).aggregate(
            total_cesantias=Sum('total_cesantias'),
            total_intereses=Sum('total_intereses_cesantias'),
            total_prima=Sum('total_prima'),
            total_vacaciones=Sum('total_vacaciones'),
            total_provisiones=Sum('total_provisiones'),
            total_salarios=Sum('total_salario_base')
        )

        meses_calculados = ResumenPrestaciones.objects.filter(año=año).count()

        return Response({
            'año': año,
            'meses_calculados': meses_calculados,
            'acumulado': acumulado
        })

    @action(detail=False, methods=['get'])
    def por_trabajador(self, request):
        """Obtener provisiones acumuladas por trabajador"""
        from datetime import date
        from django.db.models import Sum

        año = request.query_params.get('año', date.today().year)
        trabajador_id = request.query_params.get('trabajador')

        queryset = ProvisionPrestaciones.objects.filter(año=año)

        if trabajador_id:
            queryset = queryset.filter(trabajador_id=trabajador_id)

        acumulado = queryset.values(
            'trabajador__id',
            'trabajador__nombres',
            'trabajador__apellidos',
            'trabajador__numero_documento'
        ).annotate(
            total_cesantias=Sum('cesantias'),
            total_intereses=Sum('intereses_cesantias'),
            total_prima=Sum('prima'),
            total_vacaciones=Sum('vacaciones'),
            total_provisiones=Sum('total_provision'),
            total_salarios=Sum('salario_base')
        ).order_by('trabajador__apellidos', 'trabajador__nombres')

        return Response({
            'año': año,
            'trabajadores': list(acumulado)
        })
