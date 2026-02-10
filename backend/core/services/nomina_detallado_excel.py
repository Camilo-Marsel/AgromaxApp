# backend/core/services/nomina_detallado_excel.py
# Generador de Excel con reporte detallado de nómina (uso interno)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from io import BytesIO
from decimal import Decimal


class NominaDetalladoExcelGenerator:
    """Generador de Excel con detalle completo de nómina para uso interno"""

    def __init__(self, nominas, quincena):
        self.nominas = nominas
        self.quincena = quincena
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Reporte Detallado"

    def generar(self):
        """Genera el archivo Excel con reporte detallado"""

        # Configurar estilos
        self._configurar_estilos()

        # Título
        self._agregar_titulo()

        # Encabezados
        self._agregar_encabezados()

        # Datos
        self._agregar_datos()

        # Totales
        self._agregar_totales()

        # Ajustar anchos de columna
        self._ajustar_columnas()

        return self._generar_response()

    def _configurar_estilos(self):
        """Configurar estilos reutilizables"""
        self.header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=10)
        self.title_font = Font(bold=True, size=14, color="1F4E78")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')

    def _agregar_titulo(self):
        """Agregar título del documento"""
        self.ws.merge_cells('A1:P1')
        cell = self.ws['A1']

        # Incluir info de finca si todas las nóminas pertenecen a la misma finca
        fincas = set([n.trabajador.finca.nombre for n in self.nominas if n.trabajador and n.trabajador.finca])
        finca_texto = f" - {list(fincas)[0]}" if len(fincas) == 1 else ""

        cell.value = f"REPORTE DETALLADO DE NÓMINA - Q{self.quincena.numero} - {self.quincena.mes}/{self.quincena.año}{finca_texto}"
        cell.font = self.title_font
        cell.alignment = self.center_alignment

        # Línea en blanco
        self.ws.append([])

    def _agregar_encabezados(self):
        """Agregar encabezados de columnas"""
        headers = [
            'Trabajador',
            'Tipo Doc',
            'Documento',
            'Finca',
            'Días Trab.',
            'Incap. (#)',
            'IBC',
            'Aux. Trans.',
            'Total Deveng.',
            'Salud 4%',
            'Pensión 4%',
            'Total 8%',
            'Adelantos',
            'Otras Deduc.',
            'Primas/Liquid.',
            'Total Neto'
        ]

        row = 3
        for col, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border

    def _agregar_datos(self):
        """Agregar datos de nóminas con cálculos detallados"""
        from core.models import DetalleNomina, RegistroLabor
        from django.db.models import Sum

        row = 4

        for nomina in self.nominas:
            trabajador = nomina.trabajador

            # 1. Trabajador
            self.ws.cell(row=row, column=1, value=trabajador.nombre_completo)

            # 2. Tipo documento
            self.ws.cell(row=row, column=2, value=trabajador.get_tipo_documento_display())

            # 3. Número documento
            self.ws.cell(row=row, column=3, value=trabajador.numero_documento)

            # 4. Finca
            finca_nombre = trabajador.finca.nombre if trabajador.finca else 'N/A'
            self.ws.cell(row=row, column=4, value=finca_nombre)

            # 5. Días trabajados (contar días únicos con registros)
            dias_trabajados = RegistroLabor.objects.filter(
                trabajador=trabajador,
                quincena=self.quincena
            ).values_list('fecha', flat=True).distinct().count()
            self.ws.cell(row=row, column=5, value=dias_trabajados).alignment = self.center_alignment

            # 6. Incapacidades (días con incapacidad médica)
            dias_incapacidad = RegistroLabor.objects.filter(
                trabajador=trabajador,
                quincena=self.quincena,
                labor__nombre='Incapacidad Médica'
            ).values_list('fecha', flat=True).distinct().count()
            self.ws.cell(row=row, column=6, value=dias_incapacidad).alignment = self.center_alignment

            # 7. IBC (Total devengado - Auxilio de transporte)
            auxilio_transporte = DetalleNomina.objects.filter(
                nomina=nomina,
                concepto='AUXILIO_TRANSPORTE'
            ).first()
            auxilio_valor = auxilio_transporte.valor_total if auxilio_transporte else Decimal('0')
            ibc = nomina.total_devengado - auxilio_valor

            cell_ibc = self.ws.cell(row=row, column=7, value=float(ibc))
            cell_ibc.number_format = '$#,##0'
            cell_ibc.alignment = self.right_alignment

            # 8. Auxilio de transporte
            cell_aux = self.ws.cell(row=row, column=8, value=float(auxilio_valor))
            cell_aux.number_format = '$#,##0'
            cell_aux.alignment = self.right_alignment

            # 9. Total devengado
            cell_dev = self.ws.cell(row=row, column=9, value=float(nomina.total_devengado))
            cell_dev.number_format = '$#,##0'
            cell_dev.alignment = self.right_alignment

            # 10. Salud (4%)
            salud = DetalleNomina.objects.filter(
                nomina=nomina,
                concepto='SALUD'
            ).first()
            salud_valor = salud.valor_total if salud else Decimal('0')
            cell_salud = self.ws.cell(row=row, column=10, value=float(salud_valor))
            cell_salud.number_format = '$#,##0'
            cell_salud.alignment = self.right_alignment

            # 11. Pensión (4%)
            pension = DetalleNomina.objects.filter(
                nomina=nomina,
                concepto='PENSION'
            ).first()
            pension_valor = pension.valor_total if pension else Decimal('0')
            cell_pension = self.ws.cell(row=row, column=11, value=float(pension_valor))
            cell_pension.number_format = '$#,##0'
            cell_pension.alignment = self.right_alignment

            # 12. Total 8%
            total_8 = salud_valor + pension_valor
            cell_total8 = self.ws.cell(row=row, column=12, value=float(total_8))
            cell_total8.number_format = '$#,##0'
            cell_total8.alignment = self.right_alignment

            # 13. Adelantos (préstamos)
            adelantos = DetalleNomina.objects.filter(
                nomina=nomina,
                concepto='PRESTAMO'
            ).aggregate(total=Sum('valor_total'))['total'] or Decimal('0')
            cell_adelantos = self.ws.cell(row=row, column=13, value=float(adelantos))
            cell_adelantos.number_format = '$#,##0'
            cell_adelantos.alignment = self.right_alignment

            # 14. Otras deducciones (ajustes manuales)
            otras_deduc = nomina.deducciones_adicionales or Decimal('0')
            cell_otras = self.ws.cell(row=row, column=14, value=float(otras_deduc))
            cell_otras.number_format = '$#,##0'
            cell_otras.alignment = self.right_alignment

            # 15. Primas/Liquidación (buscar en devengos adicionales)
            primas_liquid = Decimal('0')
            if nomina.descripcion_devengos_adicionales:
                desc_lower = nomina.descripcion_devengos_adicionales.lower()
                if 'prima' in desc_lower or 'liquidaci' in desc_lower:
                    primas_liquid = nomina.devengos_adicionales or Decimal('0')

            cell_primas = self.ws.cell(row=row, column=15, value=float(primas_liquid))
            cell_primas.number_format = '$#,##0'
            cell_primas.alignment = self.right_alignment

            # 16. Total neto
            cell_neto = self.ws.cell(row=row, column=16, value=float(nomina.total_neto))
            cell_neto.number_format = '$#,##0'
            cell_neto.alignment = self.right_alignment
            cell_neto.font = Font(bold=True)

            # Aplicar bordes a toda la fila
            for col in range(1, 17):
                self.ws.cell(row=row, column=col).border = self.border

            row += 1

    def _agregar_totales(self):
        """Agregar fila de totales"""
        from core.models import DetalleNomina

        row = self.ws.max_row + 1

        # Celda de "TOTALES"
        total_cell = self.ws.cell(row=row, column=6)
        total_cell.value = "TOTALES:"
        total_cell.font = Font(bold=True)
        total_cell.alignment = self.right_alignment

        # Calcular totales
        total_ibc = Decimal('0')
        total_aux = Decimal('0')
        total_devengado = Decimal('0')
        total_salud = Decimal('0')
        total_pension = Decimal('0')
        total_8 = Decimal('0')
        total_adelantos = Decimal('0')
        total_otras_deduc = Decimal('0')
        total_primas = Decimal('0')
        total_neto = Decimal('0')

        for nomina in self.nominas:
            # Auxilio
            auxilio = DetalleNomina.objects.filter(
                nomina=nomina,
                concepto='AUXILIO_TRANSPORTE'
            ).first()
            aux_valor = auxilio.valor_total if auxilio else Decimal('0')
            total_aux += aux_valor

            # IBC
            ibc = nomina.total_devengado - aux_valor
            total_ibc += ibc

            # Devengado
            total_devengado += nomina.total_devengado

            # Salud
            salud = DetalleNomina.objects.filter(nomina=nomina, concepto='SALUD').first()
            salud_valor = salud.valor_total if salud else Decimal('0')
            total_salud += salud_valor

            # Pensión
            pension = DetalleNomina.objects.filter(nomina=nomina, concepto='PENSION').first()
            pension_valor = pension.valor_total if pension else Decimal('0')
            total_pension += pension_valor

            # Adelantos
            adelantos = DetalleNomina.objects.filter(
                nomina=nomina,
                concepto='PRESTAMO'
            ).aggregate(total=Sum('valor_total'))['total'] or Decimal('0')
            total_adelantos += adelantos

            # Otras deducciones
            total_otras_deduc += (nomina.deducciones_adicionales or Decimal('0'))

            # Primas/Liquidación
            if nomina.descripcion_devengos_adicionales:
                desc_lower = nomina.descripcion_devengos_adicionales.lower()
                if 'prima' in desc_lower or 'liquidaci' in desc_lower:
                    total_primas += (nomina.devengos_adicionales or Decimal('0'))

            # Neto
            total_neto += nomina.total_neto

        total_8 = total_salud + total_pension

        # Escribir totales
        fill_gray = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        # IBC (col 7)
        cell = self.ws.cell(row=row, column=7, value=float(total_ibc))
        cell.number_format = '$#,##0'
        cell.font = Font(bold=True)
        cell.alignment = self.right_alignment
        cell.fill = fill_gray
        cell.border = self.border

        # Aux (col 8)
        cell = self.ws.cell(row=row, column=8, value=float(total_aux))
        cell.number_format = '$#,##0'
        cell.font = Font(bold=True)
        cell.alignment = self.right_alignment
        cell.fill = fill_gray
        cell.border = self.border

        # Total Devengado (col 9)
        cell = self.ws.cell(row=row, column=9, value=float(total_devengado))
        cell.number_format = '$#,##0'
        cell.font = Font(bold=True)
        cell.alignment = self.right_alignment
        cell.fill = fill_gray
        cell.border = self.border

        # Salud (col 10)
        cell = self.ws.cell(row=row, column=10, value=float(total_salud))
        cell.number_format = '$#,##0'
        cell.font = Font(bold=True)
        cell.alignment = self.right_alignment
        cell.fill = fill_gray
        cell.border = self.border

        # Pensión (col 11)
        cell = self.ws.cell(row=row, column=11, value=float(total_pension))
        cell.number_format = '$#,##0'
        cell.font = Font(bold=True)
        cell.alignment = self.right_alignment
        cell.fill = fill_gray
        cell.border = self.border

        # Total 8% (col 12)
        cell = self.ws.cell(row=row, column=12, value=float(total_8))
        cell.number_format = '$#,##0'
        cell.font = Font(bold=True)
        cell.alignment = self.right_alignment
        cell.fill = fill_gray
        cell.border = self.border

        # Adelantos (col 13)
        cell = self.ws.cell(row=row, column=13, value=float(total_adelantos))
        cell.number_format = '$#,##0'
        cell.font = Font(bold=True)
        cell.alignment = self.right_alignment
        cell.fill = fill_gray
        cell.border = self.border

        # Otras Deduc (col 14)
        cell = self.ws.cell(row=row, column=14, value=float(total_otras_deduc))
        cell.number_format = '$#,##0'
        cell.font = Font(bold=True)
        cell.alignment = self.right_alignment
        cell.fill = fill_gray
        cell.border = self.border

        # Primas (col 15)
        cell = self.ws.cell(row=row, column=15, value=float(total_primas))
        cell.number_format = '$#,##0'
        cell.font = Font(bold=True)
        cell.alignment = self.right_alignment
        cell.fill = fill_gray
        cell.border = self.border

        # Total Neto (col 16)
        cell = self.ws.cell(row=row, column=16, value=float(total_neto))
        cell.number_format = '$#,##0'
        cell.font = Font(bold=True)
        cell.alignment = self.right_alignment
        cell.fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = self.border

    def _ajustar_columnas(self):
        """Ajustar ancho de columnas"""
        anchos = {
            'A': 30,  # Trabajador
            'B': 10,  # Tipo Doc
            'C': 15,  # Documento
            'D': 18,  # Finca
            'E': 12,  # Días Trab
            'F': 12,  # Incap
            'G': 15,  # IBC
            'H': 15,  # Aux Trans
            'I': 15,  # Total Deveng
            'J': 12,  # Salud
            'K': 12,  # Pensión
            'L': 12,  # Total 8%
            'M': 13,  # Adelantos
            'N': 13,  # Otras Deduc
            'O': 15,  # Primas/Liquid
            'P': 15,  # Total Neto
        }

        for col, ancho in anchos.items():
            self.ws.column_dimensions[col].width = ancho

    def _generar_response(self):
        """Generar HttpResponse con el archivo"""
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)

        # Generar nombre descriptivo del archivo
        meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        mes_nombre = meses[self.quincena.mes - 1]

        # Determinar rango de fechas de la quincena
        if self.quincena.numero == 1:
            rango = f"1-15_{mes_nombre}"
        else:
            rango = f"16-{self.quincena.fecha_fin.day}_{mes_nombre}"

        # Obtener fincas únicas de las nóminas
        fincas = set([n.trabajador.finca.nombre for n in self.nominas if n.trabajador and n.trabajador.finca])

        # Construir nombre del archivo
        if len(fincas) == 1:
            # Una sola finca
            finca_nombre = list(fincas)[0].replace(' ', '_')
            filename = f"Detallado_{finca_nombre}_{rango}_{self.quincena.año}.xlsx"
        elif len(fincas) > 1:
            # Múltiples fincas
            filename = f"Detallado_MultiFinca_{rango}_{self.quincena.año}.xlsx"
        else:
            # Sin fincas
            filename = f"Detallado_{rango}_{self.quincena.año}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response
