# backend/core/services/excel_generator.py

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO


class ExcelGenerator:
    """Generador de archivos Excel para lista de pagos"""
    
    def __init__(self, nominas, quincena):
        self.nominas = nominas
        self.quincena = quincena
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Lista de Pagos"
    
    def generar_lista_pagos(self):
        """Genera el archivo Excel con la lista de pagos"""
        
        # Configurar estilos
        self._configurar_estilos()
        
        # Título
        self._agregar_titulo()
        
        # Encabezados
        self._agregar_encabezados()
        
        # Datos
        self._agregar_datos()
        
        # Totales
        self._agregar_totales()
        
        # Ajustar anchos de columna
        self._ajustar_columnas()
        
        return self._generar_response()
    
    def _configurar_estilos(self):
        """Configurar estilos reutilizables"""
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=14, color="1F4E78")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')
    
    def _agregar_titulo(self):
        """Agregar título del documento"""
        self.ws.merge_cells('A1:G1')
        cell = self.ws['A1']
        # Incluir info de finca si todas las nóminas pertenecen a la misma finca
        fincas = set([n.trabajador.finca.nombre for n in self.nominas if n.trabajador and n.trabajador.finca])
        finca_texto = f" - {list(fincas)[0]}" if len(fincas) == 1 else ""

        cell.value = f"LISTA DE PAGOS - QUINCENA {self.quincena.numero} - {self.quincena.mes}/{self.quincena.año}{finca_texto}"
        cell.font = self.title_font
        cell.alignment = self.center_alignment
        
        # Línea en blanco
        self.ws.append([])
    
    def _agregar_encabezados(self):
        """Agregar encabezados de columnas"""
        headers = [
            'Nombre Completo',
            'Tipo Documento',
            'Número Documento',
            'Banco',
            'Tipo Cuenta',
            'Número Cuenta',
            'Valor a Pagar'
        ]
        
        row = 3
        for col, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
    
    def _agregar_datos(self):
        """Agregar datos de nóminas"""
        row = 4
        
        for nomina in self.nominas:
            trabajador = nomina.trabajador
            
            # Nombre completo
            self.ws.cell(row=row, column=1, value=trabajador.nombre_completo)
            
            # Tipo documento
            self.ws.cell(row=row, column=2, value=trabajador.get_tipo_documento_display())
            
            # Número documento
            self.ws.cell(row=row, column=3, value=trabajador.numero_documento)
            
            # Banco
            self.ws.cell(row=row, column=4, value=trabajador.banco or 'N/A')
            
            # Tipo cuenta
            tipo_cuenta = trabajador.get_tipo_cuenta_bancaria_display() if trabajador.tipo_cuenta_bancaria else 'N/A'
            self.ws.cell(row=row, column=5, value=tipo_cuenta)
            
            # Número cuenta
            self.ws.cell(row=row, column=6, value=trabajador.numero_cuenta_bancaria or 'N/A')
            
            # Valor neto
            cell_valor = self.ws.cell(row=row, column=7, value=float(nomina.total_neto))
            cell_valor.number_format = '$#,##0'
            cell_valor.alignment = self.right_alignment
            
            # Aplicar bordes a toda la fila
            for col in range(1, 8):
                self.ws.cell(row=row, column=col).border = self.border
            
            row += 1
    
    def _agregar_totales(self):
        """Agregar fila de totales"""
        row = self.ws.max_row + 1
        
        # Celda de "TOTAL"
        total_cell = self.ws.cell(row=row, column=6)
        total_cell.value = "TOTAL:"
        total_cell.font = Font(bold=True)
        total_cell.alignment = self.right_alignment
        
        # Calcular total
        total = sum(nomina.total_neto for nomina in self.nominas)
        valor_cell = self.ws.cell(row=row, column=7, value=float(total))
        valor_cell.number_format = '$#,##0'
        valor_cell.font = Font(bold=True)
        valor_cell.alignment = self.right_alignment
        valor_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Bordes
        for col in range(6, 8):
            self.ws.cell(row=row, column=col).border = self.border
    
    def _ajustar_columnas(self):
        """Ajustar ancho de columnas"""
        anchos = {
            'A': 35,  # Nombre
            'B': 15,  # Tipo Doc
            'C': 18,  # Número Doc
            'D': 20,  # Banco
            'E': 18,  # Tipo Cuenta
            'F': 20,  # Número Cuenta
            'G': 15,  # Valor
        }
        
        for col, ancho in anchos.items():
            self.ws.column_dimensions[col].width = ancho
    
    def _generar_response(self):
        """Generar HttpResponse con el archivo"""
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        filename = f"Lista_Pagos_Q{self.quincena.numero}_{self.quincena.mes}_{self.quincena.año}.xlsx"
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response